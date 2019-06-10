
import urllib.request as url_req__
from bs4 import BeautifulSoup as bsoup__
import pandas as pd__
from openpyxl import Workbook as wb__
import openpyxl.utils.dataframe as odf__
from openpyxl import load_workbook


url_of_match_ = 'http://www.espncricinfo.com/series/8039/game/1144496/australia-vs-india-14th-match-icc-cricket-world-cup-2019'
url_main_domain_ = 'http://www.espncricinfo.com/'
url_resp_ = url_req__.urlopen(url_of_match_)
PARSER_ = 'html.parser'

excel_file_loc_ = 'E:/D11/D11/'

bsoup_ = bsoup__(url_resp_, PARSER_)

h4_tag_ = bsoup_.find('h4')
a_tag_ = h4_tag_.find('a')
ground_name_ = a_tag_.string
url_ground_ = a_tag_['href']

excel_file_ = excel_file_loc_ + ground_name_ + ' records.xlsx'

url_resp_ = url_req__.urlopen(url_ground_)
bsoup_ = bsoup__(url_resp_, PARSER_)
div_tag_ = bsoup_.find('div', {'id': 'recs'})
table_tag_ = div_tag_.find('table')
tr_tags_ = table_tag_.find_all('tr', {'class': 'islast1'})
tr_tag_ = tr_tags_[1]
a_tag_ = tr_tag_.find('a')
url_ground_records_sub_domain_ = a_tag_['href']
url_ground_records_ = url_main_domain_ + url_ground_records_sub_domain_

url_resp_ = url_req__.urlopen(url_ground_records_)
bsoup_ = bsoup__(url_resp_, PARSER_)
ul_tags_ = bsoup_.find_all('ul', {'class': 'Record'})
ul_tag_batting_records_ = ul_tags_[1]
ul_tag_bowling_records_ = ul_tags_[2]
ul_tag_partnership_records_ = ul_tags_[5]

li_tags_batting_records_ = ul_tag_batting_records_.find_all('li')
li_tag_high_scores_ = li_tags_batting_records_[1]
a_tag_high_scores_ = li_tag_high_scores_.a
url_high_scores_sub_domain_ = a_tag_high_scores_['href']
url_high_scores_ = url_main_domain_ + url_high_scores_sub_domain_

df_ = pd__.read_html(url_high_scores_)[0]

wb_ = wb__()
ws_ = wb_['Sheet']
ws_.title = 'High Scores'
deployable_odf_ = odf__.dataframe_to_rows(df_, index=False, header=True)
for df_row in deployable_odf_:
    ws_.append(df_row)
wb_.save(excel_file_)

li_tags_best_bowling_ = ul_tag_bowling_records_.find_all('li')
li_tag_highest_partnership_by_runs_ = li_tags_best_bowling_[1]
a_tag_best_bowling_ = li_tag_highest_partnership_by_runs_.a
url_best_bowling_sub_domain_ = a_tag_best_bowling_['href']
url_best_bowling_ = url_main_domain_ + url_best_bowling_sub_domain_

df_ = pd__.read_html(url_best_bowling_)[0]
wb_ = load_workbook(excel_file_)
ws_ = wb_.create_sheet('Best Bowling')
deployable_odf_ = odf__.dataframe_to_rows(df_, index=False, header=True)
for df_row in deployable_odf_:
    ws_.append(df_row)
wb_.save(excel_file_)

li_tags_partnership_records_ = ul_tag_partnership_records_.find_all('li')
li_tag_highest_partnership_by_runs_ = li_tags_partnership_records_[1]
a_tag_highest_partnership_by_runs_ = li_tag_highest_partnership_by_runs_.a
url_highest_partnership_by_runs_sub_domain_ = a_tag_highest_partnership_by_runs_['href']
url_highest_partnership_by_runs_ = url_main_domain_ + url_highest_partnership_by_runs_sub_domain_

df_ = pd__.read_html(url_highest_partnership_by_runs_)[0]
wb_ = load_workbook(excel_file_)
ws_ = wb_.create_sheet('Partnerships')
deployable_odf_ = odf__.dataframe_to_rows(df_, index=False, header=True)
for df_row in deployable_odf_:
    ws_.append(df_row)
wb_.save(excel_file_)

