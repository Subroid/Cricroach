
from  bs4 import BeautifulSoup as bsoup
import urllib.request as urlreq
import openpyxl as xl
import re
import pandas as pd
from openpyxl.utils import dataframe as xldf


"""
go to url of the series then get team squads url manually
fetch player info and their ids and save 
"""

url = urlreq.urlopen("http://www.espncricinfo.com/ci/content/squad/1181315.html")
mSquadName = "South Africa"
mNewExcelFile = False
mExcelFile = "C:/Users/ASUS/Desktop/Next Match Squad.xlsx"

soup = bsoup(url, 'html.parser')
atags = soup.find_all('a')
playerTags = list()
playerNames = list()
playerIds = list()
playersDict = dict()

for atag in atags:
    atagStr = str(atag)
    if(str.find(atagStr, 'player')==21) :
        playerTags.append(atagStr)

for i in range(len(playerTags)):
    if(i%2!=0):
        playerTag = playerTags[i]

        soup2 = bsoup(playerTag, 'html.parser')
        playerNameTag = soup2.find(text=True)
        playerName = str(playerNameTag).strip()
        playerNames.append(playerName)

        playerTagStr = str(playerTag)
        playerIdList = re.findall('\d+', playerTagStr)
        playerId = playerIdList[0]
        playerIds.append(playerId)

playersDF = pd.DataFrame({'Player Name': playerNames, 'Player Id': playerIds})
print(playersDF)

if(mNewExcelFile):
    wb = xl.Workbook()
    ws = wb.create_sheet(title=mSquadName)
    ws.append(playersDF)
    wb.save(mExcelFile)
else:
    wb = xl.load_workbook(mExcelFile)
    ws = wb.create_sheet(title=mSquadName)
    excel_deployable_df = xldf.dataframe_to_rows(playersDF, index=False, header=True)

    for r in excel_deployable_df:
        ws.append(r)

    wb.save(mExcelFile)
