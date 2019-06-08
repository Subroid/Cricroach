
class StatsFetcher:

    import urllib.request as Ureq
    import pandas as Pnds
    from pandas import DataFrame as DFrame
    from openpyxl import Workbook as Wb
    from openpyxl.utils import dataframe as odf__
    import re as re__

    URL_BATSMAN_CAREER_OVERALL_STATS_ = \
        'http://stats.espncricinfo.com/ci/engine/player/253802.html?class=2;template=results;type=batting'

    URL_BATSMAN_DISMISSALS_OVERALL_STATS_ = \
        'http://stats.espncricinfo.com/ci/engine/player/253802.html?class=2;template=results;type=batting;view=dismissal_summary'
    URL_BATSMAN_DISMISSALS_BAT_1ST_STATS_ = \
        'http://stats.espncricinfo.com/ci/engine/player/253802.html?batting_fielding_first=1;class=2;filter=advanced;orderby=default;template=results;type=batting;view=dismissal_summary'
    URL_BATSMAN_DISMISSALS_BAT_2ND_STATS_ = \
        'http://stats.espncricinfo.com/ci/engine/player/253802.html?batting_fielding_first=2;class=2;filter=advanced;orderby=default;template=results;type=batting;view=dismissal_summary'

    URL_BATSMAN_INNINGS_OVERALL_STATS_ = \
        'http://stats.espncricinfo.com/ci/engine/player/253802.html?class=2;filter=advanced;orderby=default;template=results;type=batting;view=innings'
    URL_BATSMAN_INNINGS_BAT_1ST_STATS_ = \
        'http://stats.espncricinfo.com/ci/engine/player/253802.html?batting_fielding_first=1;class=2;filter=advanced;orderby=start;template=results;type=batting;view=innings'
    URL_BATSMAN_INNINGS_BAT_2ND_STATS_ = \
            'http://stats.espncricinfo.com/ci/engine/player/253802.html?batting_fielding_first=2;class=2;filter=advanced;orderby=start;template=results;type=batting;view=innings'

    #todo position based(equal to 3) http://stats.espncricinfo.com/ci/engine/player/253802.html?batting_positionmax1=3;batting_positionmin1=3;batting_positionval1=batting_position;class=2;filter=advanced;orderby=default;template=results;type=batting


    def __init__(self):
        self

    def full_fetch(self, match_url, folder_loc):

        df_list_batting_career_summary_stats_ = self.__fetch_career_summaries(self.URL_BATSMAN_CAREER_OVERALL_STATS_)
        df_batting_career_overview_stats_ = df_list_batting_career_summary_stats_[0]
        df_batting_career_summary_stats_ = df_list_batting_career_summary_stats_[1]
        df_batting_career_overview_stats_ = self.__numerize_df_stats_upto_(df_batting_career_overview_stats_, 13)
        df_batting_career_summary_stats_ = self.__numerize_df_stats_upto_(df_batting_career_summary_stats_, 13)
        df_list_batting_career_summary_stats_ = [df_batting_career_overview_stats_, df_batting_career_summary_stats_]
        self.__excelize_df_stats_(df_list_batting_career_summary_stats_, folder_loc, 'test batting career summary.xlsx')


    def __fetch_career_summaries(self, url):
        df_career_overview_stats_ = self.__fetch_career_overview(url)
        df_career_summary_stats_ = self.__fetch_career_summary(url)
        df_merged_career_overview_stats_ = \
            self.__merge_career_overview_(df_career_overview_stats_, df_career_summary_stats_)
        df_career_overview_stats_keys_lenth_ = str((len(df_career_overview_stats_.keys())))
        df_career_summary_stats_keys_lenth_ = str((len(df_career_summary_stats_.keys())))
        df_merged_career_overview_stats_ = \
            self.__clean_df_stats_(df_merged_career_overview_stats_, df_career_overview_stats_keys_lenth_)
        df_career_summary_stats_ = \
            self.__clean_df_stats_(df_career_summary_stats_, df_career_summary_stats_keys_lenth_)
        return [df_merged_career_overview_stats_, df_career_summary_stats_]

    def __read_html(self, url):
        df_list_ = self.Pnds.read_html(url)
        return df_list_

    def __fetch_career_overview(self, url):
        df_list_ = self.__read_html(url)
        df_stats_ = self.DFrame(df_list_[2])
        df_stats_ = df_stats_.set_index('Unnamed: 0')
        return df_stats_

    def __fetch_career_summary(self, url):
        df_list_ = self.__read_html(url)
        df_stats_ = self.DFrame(df_list_[3])
        df_stats_ = df_stats_.set_index('Grouping')
        return df_stats_

    def __merge_career_overview_(self, df_stats_1_, df_stats_2_):
        df_bat_1st_stats_ = df_stats_2_.loc[['matches batting first']]
        df_bat_2nd_stats_ = df_stats_2_.loc[['matches fielding first']]
        df_won_batting_1st_stats_ = df_stats_2_.loc[['won batting first']]
        df_won_batting_2nd_stats_ = df_stats_2_.loc[['won fielding first']]
        df_lost_batting_1st_stats_ = df_stats_2_.loc[['lost batting first']]
        df_lost_batting_2nd_stats_ = df_stats_2_.loc[['lost fielding first']]
        df_stats_1_ = df_stats_1_ \
            .append(df_bat_1st_stats_, sort=False) \
            .append(df_bat_2nd_stats_, sort=False) \
            .append(df_won_batting_1st_stats_, sort=False) \
            .append(df_won_batting_2nd_stats_, sort=False) \
            .append(df_lost_batting_1st_stats_, sort=False) \
            .append(df_lost_batting_2nd_stats_, sort=False)
        return df_stats_1_

    def __clean_df_stats_(self, df_stats_, keys_lenght):
        df_stats_ = df_stats_.replace(to_replace='-', value='')
        df_stats_ = df_stats_.replace(to_replace='\*', value='.01', regex=True)
        try:
            df_stats_.drop(labels='Unnamed: ' + keys_lenght, axis=1, inplace=True)
        except KeyError:
            print('error found')
        df_stats_ = df_stats_.dropna(axis=1, how='all')
        df_stats_ = df_stats_.fillna("")
        return df_stats_

    def __numerize_df_stats_upto_(self, df_stats_, upto:int):
        df_stats_.iloc[:, -upto:] = df_stats_.iloc[:, -upto:].apply(self.Pnds.to_numeric)
        return df_stats_

    def __excelize_df_stats_(self, df_list_stats_, folder_loc_, file_loc_):
        df_stats_1_ = df_list_stats_[0]
        df_stats_2_ = df_list_stats_[1]
        wb_ = self.Wb()
        excel_file_ = folder_loc_ + file_loc_
        ws_ = wb_.create_sheet('abc')
        self.__excel_ready_df_stats_(df_stats_1_, ws_)
        ws_.append([""])
        self.__excel_ready_df_stats_(df_stats_2_, ws_)

        ws_.cell(row=1, column=ws_.max_column + 2).value = '0-19'
        ws_.cell(row=1, column=ws_.max_column + 1).value = 'Last 100'
        ws_.cell(row=1, column=ws_.max_column + 1).value = 'Last 2nd 100'
        ws_.cell(row=1, column=ws_.max_column + 1).value = 'Last 50'
        ws_.cell(row=1, column=ws_.max_column + 1).value = 'Last 2nd 50'
        ws_.cell(row=1, column=ws_.max_column + 1).value = 'Last 0-19'
        ws_.cell(row=1, column=ws_.max_column + 1).value = 'Last 2nd 0-19'
        ws_.cell(row=1, column=ws_.max_column + 1).value = '4s avg'
        ws_.cell(row=1, column=ws_.max_column + 1).value = '6s avg'
        ws_.cell(row=1, column=ws_.max_column + 1).value = '100s avg'
        ws_.cell(row=1, column=ws_.max_column + 1).value = '50s avg'
        ws_.cell(row=1, column=ws_.max_column + 1).value = '0-19 avg'
        ws_.cell(row=1, column=ws_.max_column + 1).value = 'Last 5 avg'
        ws_.cell(row=1, column=ws_.max_column + 1).value = '100 due 2'
        ws_.cell(row=1, column=ws_.max_column + 1).value = '50 due 2'
        ws_.cell(row=1, column=ws_.max_column + 1).value = '0-19 due 2'
        ws_.cell(row=1, column=ws_.max_column + 1).value = '100 due'
        ws_.cell(row=1, column=ws_.max_column + 1).value = '50 due'
        ws_.cell(row=1, column=ws_.max_column + 1).value = '0-19 due'
        ws_.cell(row=1, column=ws_.max_column + 1).value = 'Last 5 avg due'

        j = 3;
        for r in ws_.iter_rows(min_row=3, max_row=ws_.max_row):

            if(j > 5):
                break

            ws_.cell(row=j, column=ws_.max_column).value = \
                '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-7).column_letter + str(j) + '-' + ws_.cell(row=j, column=ws_.max_column-28).column_letter + str(j) + ',"--")'
            ws_.cell(row=j, column=ws_.max_column-1).value = \
                 '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-8).column_letter + str(j) + '-' + ws_.cell(row=j, column=ws_.max_column-14).column_letter + str(j) + ',"--")'
            ws_.cell(row=j, column=ws_.max_column-2).value = \
                '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-9).column_letter + str(j) + '-' + ws_.cell(row=j, column=ws_.max_column-16).column_letter + str(j) + ',"--")'
            ws_.cell(row=j, column=ws_.max_column-3).value = \
                  '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-10).column_letter + str(j) + '-' + ws_.cell(row=j, column=ws_.max_column-18).column_letter + str(j) + ',"--")'
            ws_.cell(row=j, column=ws_.max_column-4).value = \
                '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-8).column_letter + str(j) + '-' + ws_.cell(row=j, column=ws_.max_column-13).column_letter + str(j) + ',"--")'
            ws_.cell(row=j, column=ws_.max_column-5).value = \
                 '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-9).column_letter + str(j) + '-' + ws_.cell(row=j, column=ws_.max_column-15).column_letter + str(j) + ',"--")'
            ws_.cell(row=j, column=ws_.max_column-6).value = \
                '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-10).column_letter + str(j) + '-' + ws_.cell(row=j, column=ws_.max_column-17).column_letter + str(j) + ',"--")'
            ws_.cell(row=j, column=ws_.max_column-8).value = \
                '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-32).column_letter + str(j) + '/' + ws_.cell(row=j, column=ws_.max_column-19).column_letter + str(j) + ',"--")'
            ws_.cell(row=j, column=ws_.max_column-9).value = \
                 '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-32).column_letter + str(j) + '/' + ws_.cell(row=j, column=ws_.max_column-24).column_letter + str(j) + ',"--")'
            ws_.cell(row=j, column=ws_.max_column-10).value = \
                '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-32).column_letter + str(j) + '/' + ws_.cell(row=j, column=ws_.max_column-25).column_letter + str(j) + ',"--")'
            ws_.cell(row=j, column=ws_.max_column-11).value = \
                  '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-27).column_letter + str(j) + '/' + ws_.cell(row=j, column=ws_.max_column-21).column_letter + str(j) + ',"--")'
            ws_.cell(row=j, column=ws_.max_column-12).value = \
                '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-27).column_letter + str(j) + '/' + ws_.cell(row=j, column=ws_.max_column-22).column_letter + str(j) + ',"--")'
            # ws_.cell(row=j, column=ws_.max_column-5).value = \
            #      '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-9).column_letter + str(j) + '-' + ws_.cell(row=j, column=ws_.max_column-15).column_letter + str(j) + ',"--")'
            # ws_.cell(row=j, column=ws_.max_column-6).value = \
            #     '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-10).column_letter + str(j) + '-' + ws_.cell(row=j, column=ws_.max_column-17).column_letter + str(j) + ',"--")'


            if(j==3):
                ws_.cell(row=j, column=ws_.max_column - 7).value = \
                    self.__find_batsman_5_matches_avg_(self.URL_BATSMAN_INNINGS_OVERALL_STATS_)
                ws_.cell(row=j, column=ws_.max_column-13).value = \
                    self.__find_batsman_2nd_last_match_(0, 20,  self.URL_BATSMAN_INNINGS_OVERALL_STATS_)
                ws_.cell(row=j, column=ws_.max_column-14).value = \
                    self.__find_batsman_last_match_(0, 20,  self.URL_BATSMAN_INNINGS_OVERALL_STATS_)
                ws_.cell(row=j, column=ws_.max_column-15).value = \
                    self.__find_batsman_2nd_last_match_(50, 100,  self.URL_BATSMAN_INNINGS_OVERALL_STATS_)
                ws_.cell(row=j, column=ws_.max_column-16).value = \
                    self.__find_batsman_last_match_(50, 100,  self.URL_BATSMAN_INNINGS_OVERALL_STATS_)
                ws_.cell(row=j, column=ws_.max_column-17).value = \
                    self.__find_batsman_2nd_last_match_(100, 300,  self.URL_BATSMAN_INNINGS_OVERALL_STATS_)
                ws_.cell(row=j, column=ws_.max_column-18).value = \
                    self.__find_batsman_last_match_(100, 300,  self.URL_BATSMAN_INNINGS_OVERALL_STATS_)
                ws_.cell(row=j, column=ws_.max_column-19).value = \
                    self.__get_bastman_dismissals_0to19_(self.URL_BATSMAN_DISMISSALS_OVERALL_STATS_)


            if(j==4):
                ws_.cell(row=j, column=ws_.max_column - 7).value = \
                    self.__find_batsman_5_matches_avg_(self.URL_BATSMAN_INNINGS_BAT_1ST_STATS_)
                ws_.cell(row=j, column=ws_.max_column-13).value = \
                    self.__find_batsman_2nd_last_match_(0, 20,  self.URL_BATSMAN_INNINGS_BAT_1ST_STATS_)
                ws_.cell(row=j, column=ws_.max_column-14).value = \
                    self.__find_batsman_last_match_(0, 20,  self.URL_BATSMAN_INNINGS_BAT_1ST_STATS_)
                ws_.cell(row=j, column=ws_.max_column-15).value = \
                    self.__find_batsman_2nd_last_match_(50, 100,  self.URL_BATSMAN_INNINGS_BAT_1ST_STATS_)
                ws_.cell(row=j, column=ws_.max_column-16).value = \
                    self.__find_batsman_last_match_(50, 100,  self.URL_BATSMAN_INNINGS_BAT_1ST_STATS_)
                ws_.cell(row=j, column=ws_.max_column-17).value = \
                    self.__find_batsman_2nd_last_match_(100, 300,  self.URL_BATSMAN_INNINGS_BAT_1ST_STATS_)
                ws_.cell(row=j, column=ws_.max_column-18).value = \
                    self.__find_batsman_last_match_(100, 300,  self.URL_BATSMAN_INNINGS_BAT_1ST_STATS_)
                ws_.cell(row=j, column=ws_.max_column-19).value = \
                    self.__get_bastman_dismissals_0to19_(self.URL_BATSMAN_DISMISSALS_BAT_1ST_STATS_)

            if(j==5):
                ws_.cell(row=j, column=ws_.max_column - 7).value = \
                    self.__find_batsman_5_matches_avg_(self.URL_BATSMAN_INNINGS_BAT_2ND_STATS_)
                ws_.cell(row=j, column=ws_.max_column-13).value = \
                    self.__find_batsman_2nd_last_match_(0, 20,  self.URL_BATSMAN_INNINGS_BAT_2ND_STATS_)
                ws_.cell(row=j, column=ws_.max_column-14).value = \
                    self.__find_batsman_last_match_(0, 20,  self.URL_BATSMAN_INNINGS_BAT_2ND_STATS_)
                ws_.cell(row=j, column=ws_.max_column-15).value = \
                    self.__find_batsman_2nd_last_match_(50, 100,  self.URL_BATSMAN_INNINGS_BAT_2ND_STATS_)
                ws_.cell(row=j, column=ws_.max_column-16).value = \
                    self.__find_batsman_last_match_(50, 100,  self.URL_BATSMAN_INNINGS_BAT_2ND_STATS_)
                ws_.cell(row=j, column=ws_.max_column-17).value = \
                    self.__find_batsman_2nd_last_match_(100, 300,  self.URL_BATSMAN_INNINGS_BAT_2ND_STATS_)
                ws_.cell(row=j, column=ws_.max_column-18).value = \
                    self.__find_batsman_last_match_(100, 300,  self.URL_BATSMAN_INNINGS_BAT_2ND_STATS_)
                ws_.cell(row=j, column=ws_.max_column-19).value = \
                    self.__get_bastman_dismissals_0to19_(self.URL_BATSMAN_DISMISSALS_BAT_2ND_STATS_)

            j = j + 1

        ws_['A1'].value = "Player Role"
        ws_['A11'].value = "Player Role"


        sheet_ = wb_['Sheet']
        wb_.remove(sheet_)
        wb_.save(excel_file_)
        wb_.close()

    def __excel_ready_df_stats_(self, df_stats_, ws_):
        excel_ready_df_stats_ = self.odf__.dataframe_to_rows(df_stats_, index=True, header=True)
        for df_row in excel_ready_df_stats_:
            ws_.append(df_row)

    def __get_bastman_dismissals_0to19_(self, url):
        num_of_dismissals_ = 0
        df_list_ = self.Pnds.read_html(url)
        df_stats_ = df_list_[3]
        df_stats_ = df_stats_.set_index('Grouping')
        runs_0_dismissals_ = self.__get_batsman_dismissals_('0 runs', df_stats_)
        runs_1to9_dismissals_ = self.__get_batsman_dismissals_('1-9 runs', df_stats_)
        runs_10to19_dismissals_ = self.__get_batsman_dismissals_('10-19 runs', df_stats_)
        num_of_dismissals_ = runs_0_dismissals_ + runs_1to9_dismissals_ + runs_10to19_dismissals_
        return num_of_dismissals_

    def __get_batsman_dismissals_(self, runs_range, df_stats):
        num_of_dismissals_ =  0
        try:
            num_of_dismissals_ = df_stats.at[runs_range, 'Dis']
        except KeyError as e:
            pass
        return num_of_dismissals_

    def __find_batsman_last_match_(self, greater_than, less_than, batsman_innings_url):
        df_list_ = self.Pnds.read_html(batsman_innings_url)
        df_stats_ = self.DFrame(df_list_[3])
        # todo * to 0.01 and and DNB to 0.00 in another method like last 5 avg
        df_stats_total_rows_ = len(df_stats_.index)
        j = 0
        found_last_match = False
        for i in range(df_stats_total_rows_-1):
            val = df_stats_.at[(df_stats_total_rows_-1) - i, 'Runs']
            if (str(val).find('*')):
                val = str(val).split('*')[0]
                try:
                    val = int(val)
                    j = j + 1
                    if (val > greater_than and val < less_than):
                        found_last_match = True
                        break
                except ValueError:
                    pass
        if(not found_last_match):
            j = '-'
        return j

    def __find_batsman_2nd_last_match_(self, greater_than, less_than, batsman_innings_url):
        df_list_ = self.Pnds.read_html(batsman_innings_url)
        df_stats_ = self.DFrame(df_list_[3])
        df_stats_total_rows_ = len(df_stats_.index)
        j = 0
        found_2nd_last_match = False
        last_match = 0
        k = 0
        for i in range(df_stats_total_rows_ - 1):
            val = df_stats_.at[(df_stats_total_rows_ - 1) - i, 'Runs']
            if (str(val).find('*')):
                val = str(val).split('*')[0]
                try:
                    val = int(val)
                    j = j + 1
                    if (val > greater_than and val < less_than):
                        last_match = last_match + 1
                        if(last_match==1):
                            k = j
                        if(last_match==2):
                            found_2nd_last_match = True
                            j = j - k
                            break
                except ValueError as e:
                    pass
        if (not found_2nd_last_match):
            j = '-'
        return j

    def __find_batsman_5_matches_avg_(self, url):
        df_list_ = self.Pnds.read_html(url)
        df_stats_ = self.DFrame(df_list_[3])
        df_stats_total_rows_ = len(df_stats_.index)
        found_last_5th_match = False
        last_match = 0
        last_5_matches_total = 0
        last_5_matches_avg = 0
        for i in range(df_stats_total_rows_ - 1):
            val = df_stats_.at[(df_stats_total_rows_ - 1) - i, 'Runs']
            try:
                int(val)
                last_match = last_match + 1
            except:
                pass
            if (str(val).find('*')):
                val = str(val).split('*')[0]
            try:
                val = int(val)
                last_5_matches_total = last_5_matches_total + val
                if (last_match == 5):
                    found_last_5th_match = True
                    last_5_matches_avg = last_5_matches_total / 5
                    break
            except ValueError as e:
                pass
        if (not found_last_5th_match):
            last_5_matches_avg = '-'
        return last_5_matches_avg