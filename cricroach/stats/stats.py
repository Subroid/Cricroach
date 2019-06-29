
class StatsFetcher:


    #todo ??bowler upto 7th bat pos dismissals ratio/avg etc


    import urllib.request as Ureq
    import pandas as Pnds
    from pandas import DataFrame as DFrame
    from openpyxl import Workbook as Wb
    from openpyxl.utils import dataframe as Odf
    from bs4 import BeautifulSoup as Bsoup
    import re as Re
    import openpyxl.styles as Ostyles

    #todo ??partnershiping
    #todo top order based(equal to 3) http://stats.espncricinfo.com/ci/engine/player/253802.html?batting_positionmax1=3;batting_positionmin1=3;batting_positionval1=batting_position;class=2;filter=advanced;orderby=default;template=results;type=batting
    #todo batting/bowling style
    # todo opener position dismissal summary http://stats.espncricinfo.com/ci/engine/player/34102.html?batting_positionmax1=2;batting_positionval1=batting_position;class=2;filter=advanced;orderby=start;orderbyad=reverse;template=results;type=batting;view=dismissal_summary
    # todo middle order http://stats.espncricinfo.com/ci/engine/player/28081.html?batting_positionmax1=7;batting_positionmin1=4;batting_positionval1=batting_position;class=2;filter=advanced;orderby=default;template=results;type=batting
    def __init__(self):
        self

    def full_fetch(self, match_url, folder_loc):

        url_resp_ = self.Ureq.urlopen(match_url)
        PARSER_ = 'html.parser'
        bsoup_ = self.Bsoup(url_resp_, PARSER_)

        team_names_span_tags_ = bsoup_.find_all('span', attrs={'class': 'team-name-short'})
        team_1_name_ = team_names_span_tags_[0].string
        team_2_name_ = team_names_span_tags_[1].string

        team_table_tags_ = bsoup_.find_all('table')
        team_1_table_tag_ = team_table_tags_[0]
        team_2_table_tag_ = team_table_tags_[1]

        team_1_tbody_tag_ = team_1_table_tag_.tbody
        team_1_tr_tags_ = team_1_tbody_tag_.find_all('tr')

        team_2_tbody_tag_ = team_2_table_tag_.tbody
        team_2_tr_tags_ = team_2_tbody_tag_.find_all('tr')

       # TEAM 1 BATTING
        wb_ = self.Wb()
        excel_file_team_1_batsmen_ = folder_loc + team_1_name_ + ' Batting career summary.xlsx'
        self.__excelize_all_batsmen_batting_analysis_(team_1_tr_tags_, wb_)
        sheet_ = wb_['Sheet']
        wb_.remove(sheet_)
        wb_.save(excel_file_team_1_batsmen_)
        wb_.close()

        # TEAM 2 BATTING
        # wb_ = self.Wb()
        # excel_file_team_2_batsmen_ = folder_loc + team_2_name_ + ' Batting career summary.xlsx'
        # self.__excelize_all_batsmen_batting_analysis_(team_2_tr_tags_, wb_)
        # sheet_ = wb_['Sheet']
        # wb_.remove(sheet_)
        # wb_.save(excel_file_team_2_batsmen_)
        # wb_.close()
        #
        # # TEAM 1 BOWLING
        # wb_ = self.Wb()
        # excel_file_team_1_bowlers_ = folder_loc + team_1_name_ + ' Bowling career summary.xlsx'
        # self.__excelize_all_bowlers_bowling_analysis_(team_1_tr_tags_, wb_)
        # sheet_ = wb_['Sheet']
        # wb_.remove(sheet_)
        # wb_.save(excel_file_team_1_bowlers_)
        # wb_.close()
        #
        # # TEAM 2 BOWLING
        # wb_ = self.Wb()
        # excel_file_team_2_bowlers_ = folder_loc + team_2_name_ + ' Bowling career summary.xlsx'
        # self.__excelize_all_bowlers_bowling_analysis_(team_2_tr_tags_, wb_)
        # sheet_ = wb_['Sheet']
        # wb_.remove(sheet_)
        # wb_.save(excel_file_team_2_bowlers_)
        # wb_.close()


    def __read_html(self, url):
        df_list_ = self.Pnds.read_html(url)
        return df_list_

    def __excelize_all_batsmen_batting_analysis_(self, team_tr_tags, wb):

        # for i in range(len(team_tr_tags)):
        for i in range(1):
            player_tr_tag_ = team_tr_tags[0]
            player_a_tag_ = player_tr_tag_.a
            player_profile_link_ = player_a_tag_['href']
            player_id_ = self.Re.findall('\d+', player_profile_link_)[0]
            player_name_ = str(player_a_tag_.string).strip()
            player_role_td_tag_ = player_tr_tag_.find('td', {'class': 'role'})
            player_role_ = str(player_role_td_tag_.string)
            print(player_name_)
            print(player_role_)
            if ('bowler' in str(player_role_).lower()):
                continue
            URL_BATSMAN_CAREER_OVERALL_STATS_ = \
                'http://stats.espncricinfo.com/ci/engine/player/'+player_id_+'.html?class=2;template=results;type=batting'

            try:
                df_list_batting_career_summary_stats_ = self.__fetch_career_summaries(URL_BATSMAN_CAREER_OVERALL_STATS_)
            except KeyError:
                continue
            df_batting_career_overview_stats_ = df_list_batting_career_summary_stats_[0]
            df_batting_career_summary_stats_ = df_list_batting_career_summary_stats_[1]
            df_batting_career_overview_stats_ = self.__numerize_df_stats_upto_(df_batting_career_overview_stats_, 13)
            df_batting_career_summary_stats_ = self.__numerize_df_stats_upto_(df_batting_career_summary_stats_, 13)
            df_list_batting_career_summary_stats_ = [df_batting_career_overview_stats_, df_batting_career_summary_stats_]
            self.__excelize_batsman_df_stats_(df_list_batting_career_summary_stats_, wb, player_name_, player_role_, player_id_)

    def __fetch_career_summaries(self, url):
        try:
            df_career_overview_stats_ = self.__fetch_career_overview(url)
        except KeyError as e:
            return e
        df_career_summary_stats_ = self.__fetch_career_summary(url)
        df_merged_career_overview_stats_ = \
            self.__merge_career_overview_(df_career_overview_stats_, df_career_summary_stats_)
        df_career_overview_stats_keys_lenth_ = str((len(df_career_overview_stats_.keys())))
        df_career_summary_stats_keys_lenth_ = str((len(df_career_summary_stats_.keys())))
        df_merged_career_overview_stats_ = \
            self.__clean_df_stats_(df_merged_career_overview_stats_, df_career_overview_stats_keys_lenth_)
        df_career_summary_stats_ = \
            self.__clean_df_stats_(df_career_summary_stats_, df_career_summary_stats_keys_lenth_)
        return [df_merged_career_overview_stats_, df_career_summary_stats_]

    def __fetch_career_overview(self, url):
        df_list_ = self.__read_html(url)
        df_stats_ = self.DFrame(df_list_[2])
        try:
         df_stats_ = df_stats_.set_index('Unnamed: 0')
        except KeyError as e:
            return e
        return df_stats_

    def __fetch_career_summary(self, url):
        df_list_ = self.__read_html(url)
        df_stats_ = self.DFrame(df_list_[3])
        df_stats_ = df_stats_.set_index('Grouping')
        return df_stats_

    def __merge_career_overview_(self, df_stats_1_, df_stats_2_):
        try:
            df_bat_1st_stats_ = df_stats_2_.loc[['matches batting first']]
        except KeyError:
            df_bat_1st_stats_ = self.DFrame()
        try:
            df_bat_2nd_stats_ = df_stats_2_.loc[['matches fielding first']]
        except KeyError:
            df_bat_2nd_stats_ = self.DFrame()
        try:
            df_won_batting_1st_stats_ = df_stats_2_.loc[['won batting first']]
        except KeyError:
            df_won_batting_1st_stats_ = self.DFrame()
        try:
            df_won_batting_2nd_stats_ = df_stats_2_.loc[['won fielding first']]
        except KeyError:
            df_won_batting_2nd_stats_ = self.DFrame()
        try:
            df_lost_batting_1st_stats_ = df_stats_2_.loc[['lost batting first']]
        except KeyError:
            df_lost_batting_1st_stats_ = self.DFrame()
        try:
            df_lost_batting_2nd_stats_ = df_stats_2_.loc[['lost fielding first']]
        except KeyError:
            df_lost_batting_2nd_stats_ = self.DFrame()
        df_stats_1_ = df_stats_1_ \
            .append(df_bat_1st_stats_, sort=False) \
            .append(df_bat_2nd_stats_, sort=False) \
            .append(df_won_batting_1st_stats_, sort=False) \
            .append(df_won_batting_2nd_stats_, sort=False) \
            .append(df_lost_batting_1st_stats_, sort=False) \
            .append(df_lost_batting_2nd_stats_, sort=False)
        return df_stats_1_

    def __clean_df_stats_(self, df_stats_, keys_lenght):
        df_stats_ = df_stats_.replace(to_replace='-', value='')
        df_stats_ = df_stats_.replace(to_replace='\*', value='.01', regex=True)
        try:
            df_stats_.drop(labels='Unnamed: ' + keys_lenght, axis=1, inplace=True)
        except KeyError:
            print('error found')
        df_stats_ = df_stats_.dropna(axis=1, how='all')
        df_stats_ = df_stats_.fillna("")
        return df_stats_

    def __numerize_df_stats_upto_(self, df_stats_, upto:int):
        df_stats_.iloc[:, -upto:] = df_stats_.iloc[:, -upto:].apply(self.Pnds.to_numeric)
        return df_stats_

    def __excelize_batsman_df_stats_(self, df_list_stats_, wb_, player_name, player_role, player_id):

        URL_BATSMAN_OPENER_OVERALL_STATS = \
            'http://stats.espncricinfo.com/ci/engine/player/'+player_id+'.html?batting_positionmax1=2;batting_positionval1=batting_position;class=2;filter=advanced;orderby=default;template=results;type=batting'

        URL_BATSMAN_OPENER_DISMISSALS_OVERALL_STATS = \
            'http://stats.espncricinfo.com/ci/engine/player/'+player_id+'.html?batting_positionmax1=2;batting_positionval1=batting_position;class=2;filter=advanced;orderby=default;template=results;type=batting;view=dismissal_summary'
        URL_BATSMAN_OPENER_DISMISSALS_OVERALL_STATS = \
            'http://stats.espncricinfo.com/ci/engine/player/'+player_id+'.html?batting_fielding_first=1;batting_positionmax1=2;batting_positionval1=batting_position;class=2;filter=advanced;orderby=default;template=results;type=batting;view=dismissal_summary'


        URL_BATSMAN_DISMISSALS_OVERALL_STATS_ = \
            'http://stats.espncricinfo.com/ci/engine/player/'+player_id+'.html?class=2;template=results;type=batting;view=dismissal_summary'
        URL_BATSMAN_DISMISSALS_BAT_1ST_STATS_ = \
            'http://stats.espncricinfo.com/ci/engine/player/'+player_id+'.html?batting_fielding_first=1;class=2;filter=advanced;orderby=default;template=results;type=batting;view=dismissal_summary'
        URL_BATSMAN_DISMISSALS_BAT_2ND_STATS_ = \
            'http://stats.espncricinfo.com/ci/engine/player/'+player_id+'.html?batting_fielding_first=2;class=2;filter=advanced;orderby=default;template=results;type=batting;view=dismissal_summary'

        URL_BATSMAN_INNINGS_OVERALL_STATS_ = \
            'http://stats.espncricinfo.com/ci/engine/player/'+player_id+'.html?class=2;filter=advanced;orderby=default;template=results;type=batting;view=innings'
        URL_BATSMAN_INNINGS_BAT_1ST_STATS_ = \
            'http://stats.espncricinfo.com/ci/engine/player/'+player_id+'.html?batting_fielding_first=1;class=2;filter=advanced;orderby=start;template=results;type=batting;view=innings'
        URL_BATSMAN_INNINGS_BAT_2ND_STATS_ = \
            'http://stats.espncricinfo.com/ci/engine/player/'+player_id+'.html?batting_fielding_first=2;class=2;filter=advanced;orderby=start;template=results;type=batting;view=innings'

        df_stats_1_ = df_list_stats_[0]
        df_stats_2_ = df_list_stats_[1]
        recent_batting_position_ = self.__get_recent_batting_position_(URL_BATSMAN_INNINGS_OVERALL_STATS_)
        recent_batting_position_df_stats_ = self.__get_recent_batting_position_df_stats_(df_stats_2_, recent_batting_position_)
        df_stats_1_ = df_stats_1_.append(recent_batting_position_df_stats_, sort=False)

        URL_BATSMAN_POS_DISMISSALS_OVERALL_STATS_ = \
            'http://stats.espncricinfo.com/ci/engine/player/' + player_id + '.html?batting_positionmax1='+str(recent_batting_position_)+';batting_positionmin1='+str(recent_batting_position_)+';batting_positionval1=batting_position;class=2;template=results;type=batting;view=dismissal_summary'
        URL_BATSMAN_POS_INNINGS_OVERALL_STATS_ = \
            'http://stats.espncricinfo.com/ci/engine/player/'+player_id+'.html?batting_positionmax1='+str(recent_batting_position_)+';batting_positionmin1='+str(recent_batting_position_)+';batting_positionval1=batting_position;class=2;filter=advanced;orderby=default;template=results;type=batting;view=innings'
        URL_BATSMAN_POS_INNINGS_BAT_1ST_STATS_ = \
             'http://stats.espncricinfo.com/ci/engine/player/'+player_id+'.html?batting_fielding_first=1;batting_positionmax1='+str(recent_batting_position_)+';batting_positionmin1='+str(recent_batting_position_)+';batting_positionval1=batting_position;class=2;filter=advanced;orderby=default;template=results;type=batting;view=innings'
        URL_BATSMAN_POS_INNINGS_BAT_2ND_STATS_ = \
             'http://stats.espncricinfo.com/ci/engine/player/'+player_id+'.html?batting_fielding_first=2;batting_positionmax1='+str(recent_batting_position_)+';batting_positionmin1='+str(recent_batting_position_)+';batting_positionval1=batting_position;class=2;filter=advanced;orderby=default;template=results;type=batting;view=innings'
        URL_BATSMAN_POS_1_DISMISSALS_OVERALL_STATS_ = \
            'http://stats.espncricinfo.com/ci/engine/player/' + player_id + '.html?batting_positionmax1=1;batting_positionmin1=1batting_positionval1=batting_position;class=2;template=results;type=batting;view=dismissal_summary'
        URL_BATSMAN_POS_2_DISMISSALS_OVERALL_STATS_ = \
            'http://stats.espncricinfo.com/ci/engine/player/' + player_id + '.html?batting_positionmax1=2;batting_positionmin1=2batting_positionval1=batting_position;class=2;template=results;type=batting;view=dismissal_summary'
        URL_BATSMAN_POS_1_INNINGS_OVERALL_STATS_ = \
            'http://stats.espncricinfo.com/ci/engine/player/'+player_id+'.html?batting_positionmax1=1;batting_positionmin1=1;batting_positionval1=batting_position;class=2;filter=advanced;orderby=start;template=results;type=batting;view=innings'
        URL_BATSMAN_POS_2_INNINGS_OVERALL_STATS_ = \
                    'http://stats.espncricinfo.com/ci/engine/player/'+player_id+'.html?batting_positionmax1=2;batting_positionmin1=2;batting_positionval1=batting_position;class=2;filter=advanced;orderby=start;template=results;type=batting;view=innings'


        ws_ = wb_.create_sheet(player_name + ' ' + str(recent_batting_position_))
        self.__excel_ready_df_stats_(df_stats_1_, ws_)
        ws_.append([""])
        self.__excel_ready_df_stats_(df_stats_2_, ws_)

        ws_.cell(row=1, column=ws_.max_column + 2).value = '50+'
        ws_.cell(row=1, column=ws_.max_column + 1).value = 'BF avg'
        ws_.cell(row=1, column=ws_.max_column + 1).value = '50+ avg'
        ws_.cell(row=1, column=ws_.max_column + 1).value = '0-19'
        ws_.cell(row=1, column=ws_.max_column + 1).value = 'Last 100'
        ws_.cell(row=1, column=ws_.max_column + 1).value = 'Last 2nd 100'
        ws_.cell(row=1, column=ws_.max_column + 1).value = 'Last 50'
        ws_.cell(row=1, column=ws_.max_column + 1).value = 'Last 2nd 50'
        ws_.cell(row=1, column=ws_.max_column + 1).value = 'Last 0-19'
        ws_.cell(row=1, column=ws_.max_column + 1).value = 'Last 2nd 0-19'
        ws_.cell(row=1, column=ws_.max_column + 1).value = '4s avg'
        ws_.cell(row=1, column=ws_.max_column + 1).value = '6s avg'
        ws_.cell(row=1, column=ws_.max_column + 1).value = '100s avg'
        ws_.cell(row=1, column=ws_.max_column + 1).value = '50s avg'
        ws_.cell(row=1, column=ws_.max_column + 1).value = '0-19 avg'
        ws_.cell(row=1, column=ws_.max_column + 1).value = 'Last 5 avg'
        ws_.cell(row=1, column=ws_.max_column + 1).value = '100 due 2'
        ws_.cell(row=1, column=ws_.max_column + 1).value = '50 due 2'
        ws_.cell(row=1, column=ws_.max_column + 1).value = '0-19 due 2'
        ws_.cell(row=1, column=ws_.max_column + 1).value = '100 due'
        ws_.cell(row=1, column=ws_.max_column + 1).value = '50 due'
        ws_.cell(row=1, column=ws_.max_column + 1).value = '0-19 due'
        ws_.cell(row=1, column=ws_.max_column + 1).value = 'Last 5 avg due'

        j = 3;
        for r in ws_.iter_rows(min_row=3, max_row=ws_.max_row):

            if(j > 11):
                break
            if(j == 3 or j == 4 or j == 5 or j == 10):
                # Last 5 avg due
                ws_.cell(row=j, column=ws_.max_column).value = \
                    '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-7).column_letter + str(j) + '-' + ws_.cell(row=j, column=ws_.max_column-31).column_letter + str(j) + ',"--")'
                # 0-19 due
                ws_.cell(row=j, column=ws_.max_column-1).value = \
                     '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-8).column_letter + str(j) + '-' + ws_.cell(row=j, column=ws_.max_column-14).column_letter + str(j) + ',"--")'
                # 50 due
                ws_.cell(row=j, column=ws_.max_column-2).value = \
                    '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-9).column_letter + str(j) + '-' + ws_.cell(row=j, column=ws_.max_column-16).column_letter + str(j) + ',"--")'
                # 100 due
                ws_.cell(row=j, column=ws_.max_column-3).value = \
                      '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-10).column_letter + str(j) + '-' + ws_.cell(row=j, column=ws_.max_column-18).column_letter + str(j) + ',"--")'
                # 0-19 due 2
                ws_.cell(row=j, column=ws_.max_column-4).value = \
                    '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-8).column_letter + str(j) + '-' + ws_.cell(row=j, column=ws_.max_column-13).column_letter + str(j) + ',"--")'
                # 50 due 2
                ws_.cell(row=j, column=ws_.max_column-5).value = \
                     '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-9).column_letter + str(j) + '-' + ws_.cell(row=j, column=ws_.max_column-15).column_letter + str(j) + ',"--")'
                # 100 due 2
                ws_.cell(row=j, column=ws_.max_column-6).value = \
                    '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-10).column_letter + str(j) + '-' + ws_.cell(row=j, column=ws_.max_column-17).column_letter + str(j) + ',"--")'
                # 0-19 avg
                ws_.cell(row=j, column=ws_.max_column-9).value = \
                     '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-35).column_letter + str(j) + '/' + ws_.cell(row=j, column=ws_.max_column-19).column_letter + str(j) + ',"--")'
                # 50s avg
                ws_.cell(row=j, column=ws_.max_column-10).value = \
                    '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-35).column_letter + str(j) + '/' + ws_.cell(row=j, column=ws_.max_column-27).column_letter + str(j) + ',"--")'
                # 100s avg
                ws_.cell(row=j, column=ws_.max_column-11).value = \
                      '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-35).column_letter + str(j) + '/' + ws_.cell(row=j, column=ws_.max_column-28).column_letter + str(j) + ',"--")'
                # 6s avg
                ws_.cell(row=j, column=ws_.max_column-12).value = \
                    '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-24).column_letter + str(j) + '/' + ws_.cell(row=j, column=ws_.max_column-35).column_letter + str(j) + ',"--")'
                # 4s avg
                ws_.cell(row=j, column=ws_.max_column-12).value = \
                    '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-25).column_letter + str(j) + '/' + ws_.cell(row=j, column=ws_.max_column-35).column_letter + str(j) + ',"--")'
                # 50+ avg
                ws_.cell(row=j, column=ws_.max_column-20).value = \
                                    '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-35).column_letter + str(j) + '/' + ws_.cell(row=j, column=ws_.max_column-22).column_letter + str(j) + ',"--")'
                # BF avg
                ws_.cell(row=j, column=ws_.max_column-21).value = \
                                                    '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-30).column_letter + str(j) + '/' + ws_.cell(row=j, column=ws_.max_column-35).column_letter + str(j) + ',"--")'
                # 50+
                ws_.cell(row=j, column=ws_.max_column-22).value = \
                                                    '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-28).column_letter + str(j) + '+' + ws_.cell(row=j, column=ws_.max_column-27).column_letter + str(j) + ',"--")'

                ws_.cell(row=j, column=ws_.max_column - 1).number_format = '#,##0.00'
                ws_.cell(row=j, column=ws_.max_column - 2).number_format = '#,##0.00'
                ws_.cell(row=j, column=ws_.max_column - 3).number_format = '#,##0.00'
                ws_.cell(row=j, column=ws_.max_column - 4).number_format = '#,##0.00'
                ws_.cell(row=j, column=ws_.max_column - 5).number_format = '#,##0.00'
                ws_.cell(row=j, column=ws_.max_column - 6).number_format = '#,##0.00'
                ws_.cell(row=j, column=ws_.max_column - 8).number_format = '#,##0.00'
                ws_.cell(row=j, column=ws_.max_column - 9).number_format = '#,##0.00'
                ws_.cell(row=j, column=ws_.max_column - 10).number_format = '#,##0.00'
                ws_.cell(row=j, column=ws_.max_column - 11).number_format = '#,##0.00'
                ws_.cell(row=j, column=ws_.max_column - 12).number_format = '#,##0.00'
                ws_.cell(row=j, column=ws_.max_column - 13).number_format = '#,##0.00'
                ws_.cell(row=j, column=ws_.max_column - 20).number_format = '#,##0.00'
                ws_.cell(row=j, column=ws_.max_column - 21).number_format = '#,##0.00'

            if(j==3):
                # Last 5 avg
                ws_.cell(row=j, column=ws_.max_column - 8).value = \
                    self.__find_batsman_5_matches_avg_(URL_BATSMAN_INNINGS_OVERALL_STATS_)
                ws_.cell(row=j, column=ws_.max_column-13).value = \
                    self.__find_batsman_2nd_last_match_(-1, 20, URL_BATSMAN_INNINGS_OVERALL_STATS_)
                ws_.cell(row=j, column=ws_.max_column-14).value = \
                    self.__find_batsman_last_match_(-1, 20, URL_BATSMAN_INNINGS_OVERALL_STATS_)
                ws_.cell(row=j, column=ws_.max_column-15).value = \
                    self.__find_batsman_2nd_last_match_(49, 100, URL_BATSMAN_INNINGS_OVERALL_STATS_)
                ws_.cell(row=j, column=ws_.max_column-16).value = \
                    self.__find_batsman_last_match_(49, 100, URL_BATSMAN_INNINGS_OVERALL_STATS_)
                ws_.cell(row=j, column=ws_.max_column-17).value = \
                    self.__find_batsman_2nd_last_match_(99, 300, URL_BATSMAN_INNINGS_OVERALL_STATS_)
                ws_.cell(row=j, column=ws_.max_column-18).value = \
                    self.__find_batsman_last_match_(99, 300, URL_BATSMAN_INNINGS_OVERALL_STATS_)
                ws_.cell(row=j, column=ws_.max_column-19).value = \
                    self.__get_bastman_dismissals_0to19_(URL_BATSMAN_DISMISSALS_OVERALL_STATS_)


            if(j==4):
                # Last 5 avg
                ws_.cell(row=j, column=ws_.max_column - 8).value = \
                    self.__find_batsman_5_matches_avg_(URL_BATSMAN_INNINGS_OVERALL_STATS_)
                ws_.cell(row=j, column=ws_.max_column-13).value = \
                    self.__find_batsman_2nd_last_match_(-1, 20, URL_BATSMAN_INNINGS_BAT_1ST_STATS_)
                ws_.cell(row=j, column=ws_.max_column-14).value = \
                    self.__find_batsman_last_match_(-1, 20, URL_BATSMAN_INNINGS_BAT_1ST_STATS_)
                ws_.cell(row=j, column=ws_.max_column-15).value = \
                    self.__find_batsman_2nd_last_match_(49, 100, URL_BATSMAN_INNINGS_BAT_1ST_STATS_)
                ws_.cell(row=j, column=ws_.max_column-16).value = \
                    self.__find_batsman_last_match_(49, 100, URL_BATSMAN_INNINGS_BAT_1ST_STATS_)
                ws_.cell(row=j, column=ws_.max_column-17).value = \
                    self.__find_batsman_2nd_last_match_(99, 300, URL_BATSMAN_INNINGS_BAT_1ST_STATS_)
                ws_.cell(row=j, column=ws_.max_column-18).value = \
                    self.__find_batsman_last_match_(99, 300, URL_BATSMAN_INNINGS_BAT_1ST_STATS_)
                ws_.cell(row=j, column=ws_.max_column-19).value = \
                    self.__get_bastman_dismissals_0to19_(URL_BATSMAN_DISMISSALS_BAT_1ST_STATS_)

            if(j==5):
                # Last 5 avg
                ws_.cell(row=j, column=ws_.max_column - 8).value = \
                    self.__find_batsman_5_matches_avg_(URL_BATSMAN_INNINGS_OVERALL_STATS_)
                ws_.cell(row=j, column=ws_.max_column-13).value = \
                    self.__find_batsman_2nd_last_match_(-1, 20, URL_BATSMAN_INNINGS_BAT_2ND_STATS_)
                ws_.cell(row=j, column=ws_.max_column-14).value = \
                    self.__find_batsman_last_match_(-1, 20, URL_BATSMAN_INNINGS_BAT_2ND_STATS_)
                ws_.cell(row=j, column=ws_.max_column-15).value = \
                    self.__find_batsman_2nd_last_match_(49, 100, URL_BATSMAN_INNINGS_BAT_2ND_STATS_)
                ws_.cell(row=j, column=ws_.max_column-16).value = \
                    self.__find_batsman_last_match_(49, 100, URL_BATSMAN_INNINGS_BAT_2ND_STATS_)
                ws_.cell(row=j, column=ws_.max_column-17).value = \
                    self.__find_batsman_2nd_last_match_(99, 300, URL_BATSMAN_INNINGS_BAT_2ND_STATS_)
                ws_.cell(row=j, column=ws_.max_column-18).value = \
                    self.__find_batsman_last_match_(99, 300, URL_BATSMAN_INNINGS_BAT_2ND_STATS_)
                ws_.cell(row=j, column=ws_.max_column-19).value = \
                    self.__get_bastman_dismissals_0to19_(URL_BATSMAN_DISMISSALS_BAT_2ND_STATS_)

            if (j == 10):
                # Last 5 avg
                ws_.cell(row=j, column=ws_.max_column - 8).value = \
                    self.__find_batsman_5_matches_avg_(URL_BATSMAN_INNINGS_OVERALL_STATS_)
                ws_.cell(row=j, column=ws_.max_column-13).value = \
                    self.__find_batsman_2nd_last_match_(-1, 20, URL_BATSMAN_POS_INNINGS_OVERALL_STATS_)
                ws_.cell(row=j, column=ws_.max_column-14).value = \
                    self.__find_batsman_last_match_(-1, 20, URL_BATSMAN_POS_INNINGS_OVERALL_STATS_)
                ws_.cell(row=j, column=ws_.max_column-15).value = \
                    self.__find_batsman_2nd_last_match_(49, 100, URL_BATSMAN_POS_INNINGS_OVERALL_STATS_)
                ws_.cell(row=j, column=ws_.max_column-16).value = \
                    self.__find_batsman_last_match_(49, 100, URL_BATSMAN_POS_INNINGS_OVERALL_STATS_)
                ws_.cell(row=j, column=ws_.max_column-17).value = \
                    self.__find_batsman_2nd_last_match_(99, 300, URL_BATSMAN_POS_INNINGS_OVERALL_STATS_)
                ws_.cell(row=j, column=ws_.max_column-18).value = \
                    self.__find_batsman_last_match_(99, 300, URL_BATSMAN_POS_INNINGS_OVERALL_STATS_)
                ws_.cell(row=j, column=ws_.max_column-19).value = \
                    self.__get_bastman_dismissals_0to19_(URL_BATSMAN_POS_DISMISSALS_OVERALL_STATS_)


            j = j + 1

        ws_['A1'].value = player_role

        ws_.freeze_panes = ws_.cell(row=3, column=ws_.max_column-23)
        self.bolding_header_font(ws_)
        self.align_center(ws_)
        # self.adjust_column_width(ws_)

        #todo conditional formatting for Ave, Strike rate, 100s avg, 50s avg, 0-19 avg, last 5 avg, dues negative numbers

    def __get_recent_batting_position_(self, url):
        df_list_ = self.Pnds.read_html(url)
        df_stats_ = self.DFrame(df_list_[3])
        df_stats_total_rows_ = len(df_stats_.index)
        found_batting_pos = False
        for i in range(df_stats_total_rows_ - 1):
            pos = ''
            try:
                pos = df_stats_.at[(df_stats_total_rows_ - 1) - i, 'Pos']
            except ValueError:
                pass
            try:
                pos = int(pos)
                found_batting_pos = True
                break
            except ValueError:
                pass
        if (not found_batting_pos):
            pos = '-'
        return pos

#todo check last 3 matches batting position then determine for opener, top order batsman, middle order batsman
    def __get_recent_batting_position_df_stats_(self, df_stats, batting_position):
        df_stats = self.DFrame(df_stats)
        if (batting_position == 1):
            batting_position = '1st position'
        elif (batting_position == 2):
            batting_position = '2nd position'
        elif (batting_position == 3):
            batting_position = '3rd position'
        else:
            batting_position = str(batting_position) + 'th position'
        print(batting_position)
        # if ('1st' in batting_position or '2nd' in batting_position):
        #     df_1st_batting_position_stats = df_stats.loc[['1st position']]
        #     df_2nd_batting_position_stats = df_stats.loc[['2nd position']]
        #     df_recent_batting_position_stats = df_1st_batting_position_stats.append(df_2nd_batting_position_stats, sort=False)
        # else:
        df_recent_batting_position_stats = df_stats.loc[[batting_position]]
        return df_recent_batting_position_stats

    def __get_bastman_dismissals_0to19_(self, url):
        num_of_dismissals_ = 0
        df_list_ = self.Pnds.read_html(url)
        df_stats_ = self.DFrame(df_list_[3])
        try:
            df_stats_ = df_stats_.set_index('Grouping')
        except KeyError:
            try:
                df_stats_ = df_stats_.set_index('')
            except KeyError:
                pass
        runs_0_dismissals_ = self.__get_batsman_dismissals_('0 runs', df_stats_)
        runs_1to9_dismissals_ = self.__get_batsman_dismissals_('1-9 runs', df_stats_)
        runs_10to19_dismissals_ = self.__get_batsman_dismissals_('10-19 runs', df_stats_)
        num_of_dismissals_ = runs_0_dismissals_ + runs_1to9_dismissals_ + runs_10to19_dismissals_
        return num_of_dismissals_

    def __get_batsman_dismissals_(self, runs_range, df_stats):
        num_of_dismissals_ =  0
        try:
            num_of_dismissals_ = df_stats.at[runs_range, 'Dis']
        except KeyError:
            pass
        except ValueError:
            pass
        return num_of_dismissals_

    def __find_batsman_last_match_(self, greater_than, less_than, batsman_innings_url):
        df_list_ = self.Pnds.read_html(batsman_innings_url)
        df_stats_ = self.DFrame(df_list_[3])
        # todo * to 0.01 and and DNB to 0.00 in another method like last 5 avg
        df_stats_total_rows_ = len(df_stats_.index)
        j = 0
        found_last_match = False
        for i in range(df_stats_total_rows_-1):
            val = ''
            try:
                val = df_stats_.at[(df_stats_total_rows_-1) - i, 'Runs']
            except ValueError:
                pass
            if (str(val).find('*')):
                val = str(val).split('*')[0]
                try:
                    val = int(val)
                    j = i
                    if (val > greater_than and val < less_than):
                        found_last_match = True
                        break
                except ValueError:
                    pass
        if(not found_last_match):
            j = '-'
        return j

    def __find_batsman_2nd_last_match_(self, greater_than, less_than, batsman_innings_url):
        df_list_ = self.Pnds.read_html(batsman_innings_url)
        df_stats_ = self.DFrame(df_list_[3])
        df_stats_total_rows_ = len(df_stats_.index)
        j = 0
        found_2nd_last_match = False
        last_match = 0
        k = 0
        for i in range(df_stats_total_rows_ - 1):
            val = ''
            try:
                val = df_stats_.at[(df_stats_total_rows_ - 1) - i, 'Runs']
            except ValueError:
                pass
            if (str(val).find('*')):
                val = str(val).split('*')[0]
                try:
                    val = int(val)
                    j = i
                    if (val > greater_than and val < less_than):
                        last_match = last_match + 1
                        if(last_match==1):
                            k = j
                        if(last_match==2):
                            found_2nd_last_match = True
                            j = j - k
                            break
                except ValueError as e:
                    pass
        if (not found_2nd_last_match):
            j = '-'
        return j

    def __find_batsman_5_matches_avg_(self, url):
        df_list_ = self.Pnds.read_html(url)
        df_stats_ = self.DFrame(df_list_[3])
        df_stats_total_rows_ = len(df_stats_.index)
        found_last_5th_match = False
        last_match = 0
        last_5_matches_total = 0
        last_5_matches_avg = 0
        for i in range(df_stats_total_rows_ - 1):
            val = ''
            try:
                val = df_stats_.at[(df_stats_total_rows_ - 1) - i, 'Runs']
            except ValueError:
                pass
            try:
                int(val)
                last_match = last_match + 1
            except:
                pass
            if (str(val).find('*')):
                val = str(val).split('*')[0]
            try:
                val = int(val)
                last_5_matches_total = last_5_matches_total + val
                if (last_match == 5):
                    found_last_5th_match = True
                    last_5_matches_avg = last_5_matches_total / 5
                    break
            except ValueError:
                pass
        if (not found_last_5th_match):
            last_5_matches_avg = '-'
        return last_5_matches_avg


    def __excelize_all_bowlers_bowling_analysis_(self, team_tr_tags, wb):

        for i in range(len(team_tr_tags)):
            player_tr_tag_ = team_tr_tags[i]
            player_a_tag_ = player_tr_tag_.a
            player_profile_link_ = player_a_tag_['href']
            player_id_ = self.Re.findall('\d+', player_profile_link_)[0]
            player_name_ = str(player_a_tag_.string).strip()
            player_role_td_tag_ = player_tr_tag_.find('td', {'class': 'role'})
            player_role_ = str(player_role_td_tag_.string)
            print(player_name_)
            print(player_role_)
            if ('batsman' in str(player_role_).lower()):
                continue
            URL_BATSMAN_CAREER_OVERALL_STATS_ = \
                'http://stats.espncricinfo.com/ci/engine/player/'+player_id_+'.html?class=2;template=results;type=bowling'

            df_list_bowling_career_summary_stats_ = self.__fetch_career_summaries(URL_BATSMAN_CAREER_OVERALL_STATS_)
            df_bowling_career_overview_stats_ = df_list_bowling_career_summary_stats_[0]
            df_bowling_career_summary_stats_ = df_list_bowling_career_summary_stats_[1]
            # df_bowling_career_overview_stats_ = self.__numerize_df_stats_upto_(df_bowling_career_overview_stats_, 13)
            # df_bowling_career_summary_stats_ = self.__numerize_df_stats_upto_(df_bowling_career_summary_stats_, 13)
            df_list_bowling_career_summary_stats_ = [df_bowling_career_overview_stats_, df_bowling_career_summary_stats_]
            self.__excelize_bowler_df_stats_(df_list_bowling_career_summary_stats_, wb, player_name_, player_role_, player_id_)


    def __excelize_bowler_df_stats_(self, df_list_stats_, wb_, player_name, player_role, player_id):

        # http://stats.espncricinfo.com/ci/engine/player/19264.html?class=2;template=results;type=bowling;view=dismissal_summary
        URL_BOWLER_DISMISSALS_OVERALL_STATS_ = \
            'http://stats.espncricinfo.com/ci/engine/player/'+player_id+'.html?class=2;template=results;type=bowling;view=dismissal_summary'
        URL_BOWLER_DISMISSALS_BAT_1ST_STATS_ = \
            'http://stats.espncricinfo.com/ci/engine/player/'+player_id+'.html?batting_fielding_first=1;class=2;filter=advanced;orderby=default;template=results;type=bowling;view=dismissal_summary'
        URL_BOWLER_DISMISSALS_BAT_2ND_STATS_ = \
            'http://stats.espncricinfo.com/ci/engine/player/'+player_id+'.html?batting_fielding_first=2;class=2;filter=advanced;orderby=default;template=results;type=bowling`;view=dismissal_summary'

        URL_BOWLER_INNINGS_OVERALL_STATS_ = \
            'http://stats.espncricinfo.com/ci/engine/player/'+player_id+'.html?class=2;filter=advanced;orderby=default;template=results;type=bowling;view=innings'
        URL_BOWLER_INNINGS_BAT_1ST_STATS_ = \
            'http://stats.espncricinfo.com/ci/engine/player/'+player_id+'.html?batting_fielding_first=1;class=2;filter=advanced;orderby=start;template=results;type=bowling;view=innings'
        URL_BOWLER_INNINGS_BAT_2ND_STATS_ = \
            'http://stats.espncricinfo.com/ci/engine/player/'+player_id+'.html?batting_fielding_first=2;class=2;filter=advanced;orderby=start;template=results;type=bowling;view=innings'

        df_stats_1_ = df_list_stats_[0]
        df_stats_2_ = df_list_stats_[1]
        recent_bowling_position_ = self.__get_recent_bowling_position_(URL_BOWLER_INNINGS_OVERALL_STATS_)
        recent_bowling_position_df_stats_ = self.__get_recent_bowling_position_df_stats_(df_stats_2_, recent_bowling_position_)
        df_stats_1_ = df_stats_1_.append(recent_bowling_position_df_stats_, sort=False)

        URL_BOWLER_POS_INNINGS_OVERALL_STATS_ = \
            'http://stats.espncricinfo.com/ci/engine/player/' + player_id + '.html?batting_positionmax1=' + str(recent_bowling_position_) + ';batting_positionmin1=' + str(
                recent_bowling_position_) + ';batting_positionval1=batting_position;class=2;filter=advanced;orderby=default;template=results;type=bowling;view=innings'
        URL_BOWLER_POS_INNINGS_BAT_1ST_STATS_ = \
            'http://stats.espncricinfo.com/ci/engine/player/' + player_id + '.html?batting_fielding_first=1;batting_positionmax1=' + str(recent_bowling_position_) + ';batting_positionmin1=' + str(
                recent_bowling_position_) + ';batting_positionval1=batting_position;class=2;filter=advanced;orderby=default;template=results;type=bowling;view=innings'
        URL_BOWLER_POS_INNINGS_BAT_2ND_STATS_ = \
            'http://stats.espncricinfo.com/ci/engine/player/' + player_id + '.html?batting_fielding_first=2;batting_positionmax1=' + str(recent_bowling_position_) + ';batting_positionmin1=' + str(
                recent_bowling_position_) + ';batting_positionval1=batting_position;class=2;filter=advanced;orderby=default;template=results;type=bowling;view=innings'

        ws_ = wb_.create_sheet(player_name + ' ' + str(recent_bowling_position_))
        self.__excel_ready_df_stats_(df_stats_1_, ws_)
        ws_.append([""])
        self.__excel_ready_df_stats_(df_stats_2_, ws_)

        ws_.cell(row=1, column=ws_.max_column + 2).value = '4+'
        ws_.cell(row=1, column=ws_.max_column + 1).value = '2+'
        ws_.cell(row=1, column=ws_.max_column + 1).value = '0'
        ws_.cell(row=1, column=ws_.max_column + 1).value = 'Last 4+'
        ws_.cell(row=1, column=ws_.max_column + 1).value = 'Last 2nd 4+'
        ws_.cell(row=1, column=ws_.max_column + 1).value = 'Last 2+'
        ws_.cell(row=1, column=ws_.max_column + 1).value = 'Last 2nd 2+'
        ws_.cell(row=1, column=ws_.max_column + 1).value = 'Last 0'
        ws_.cell(row=1, column=ws_.max_column + 1).value = 'Last 2nd 0'
        ws_.cell(row=1, column=ws_.max_column + 1).value = 'Last inn'
        ws_.cell(row=1, column=ws_.max_column + 1).value = 'Inns avg'
        ws_.cell(row=1, column=ws_.max_column + 1).value = 'Overs avg'
        ws_.cell(row=1, column=ws_.max_column + 1).value = 'Mdns avg'
        ws_.cell(row=1, column=ws_.max_column + 1).value = '4+ avg'
        ws_.cell(row=1, column=ws_.max_column + 1).value = '2+ avg'
        ws_.cell(row=1, column=ws_.max_column + 1).value = '0 avg'
        ws_.cell(row=1, column=ws_.max_column + 1).value = 'Last 5 SR'
        ws_.cell(row=1, column=ws_.max_column + 1).value = 'Inn due'
        ws_.cell(row=1, column=ws_.max_column + 1).value = '4+ due 2'
        ws_.cell(row=1, column=ws_.max_column + 1).value = '2+ due 2'
        ws_.cell(row=1, column=ws_.max_column + 1).value = '0 due 2'
        ws_.cell(row=1, column=ws_.max_column + 1).value = '4+ due'
        ws_.cell(row=1, column=ws_.max_column + 1).value = '2+ due'
        ws_.cell(row=1, column=ws_.max_column + 1).value = '0 due'
        ws_.cell(row=1, column=ws_.max_column + 1).value = 'Last 5 due'

        j = 3;
        for r in ws_.iter_rows(min_row=3, max_row=ws_.max_row):

            if (j > 10):
                break
            if (j == 3 or j == 4 or j == 5 or j == 10):

                # Last 5 due
                ws_.cell(row=j, column=ws_.max_column).value = \
                    '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-28).column_letter + str(j) + '-' + ws_.cell(row=j, column=ws_.max_column-8).column_letter + str(j) + ',"--")'
                # 0 due
                ws_.cell(row=j, column=ws_.max_column-1).value = \
                     '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-9).column_letter + str(j) + '-' + ws_.cell(row=j, column=ws_.max_column-17).column_letter + str(j) + ',"--")'
                # 2+ due
                ws_.cell(row=j, column=ws_.max_column-2).value = \
                    '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-10).column_letter + str(j) + '-' + ws_.cell(row=j, column=ws_.max_column-19).column_letter + str(j) + ',"--")'
                # 4 + due
                ws_.cell(row=j, column=ws_.max_column-3).value = \
                      '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-11).column_letter + str(j) + '-' + ws_.cell(row=j, column=ws_.max_column-21).column_letter + str(j) + ',"--")'
                # 0 due 2
                ws_.cell(row=j, column=ws_.max_column-4).value = \
                    '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-9).column_letter + str(j) + '-' + ws_.cell(row=j, column=ws_.max_column-16).column_letter + str(j) + ',"--")'
                # 2+ due 2
                ws_.cell(row=j, column=ws_.max_column-5).value = \
                     '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-10).column_letter + str(j) + '-' + ws_.cell(row=j, column=ws_.max_column-18).column_letter + str(j) + ',"--")'
                # 4+ due 2
                ws_.cell(row=j, column=ws_.max_column-6).value = \
                    '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-11).column_letter + str(j) + '-' + ws_.cell(row=j, column=ws_.max_column-20).column_letter + str(j) + ',"--")'
                # Inn due
                ws_.cell(row=j, column=ws_.max_column-7).value = \
                    '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-14).column_letter + str(j) + '-' + ws_.cell(row=j, column=ws_.max_column-15).column_letter + str(j) + ',"--")'
                # 0 avg
                ws_.cell(row=j, column=ws_.max_column-9).value = \
                     '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-36).column_letter + str(j) + '/' + ws_.cell(row=j, column=ws_.max_column-22).column_letter + str(j) + ',"--")'
                # 2+ avg
                ws_.cell(row=j, column=ws_.max_column-10).value = \
                    '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-36).column_letter + str(j) + '/' + ws_.cell(row=j, column=ws_.max_column-23).column_letter + str(j) + ',"--")'
                # 4+ avg
                ws_.cell(row=j, column=ws_.max_column-11).value = \
                      '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-36).column_letter + str(j) + '/' + ws_.cell(row=j, column=ws_.max_column-24).column_letter + str(j) + ',"--")'
                # Mdns avg
                ws_.cell(row=j, column=ws_.max_column-12).value = \
                    '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-35).column_letter + str(j) + '/' + ws_.cell(row=j, column=ws_.max_column-34).column_letter + str(j) + ',"--")'
                # Overs avg
                ws_.cell(row=j, column=ws_.max_column-13).value = \
                     '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-35).column_letter + str(j) + '/' + ws_.cell(row=j, column=ws_.max_column-36).column_letter + str(j) + ',"--")'
                # Inns avg
                ws_.cell(row=j, column=ws_.max_column-14).value = \
                    '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-37).column_letter + str(j) + '/' + ws_.cell(row=j, column=ws_.max_column-36).column_letter + str(j) + ',"--")'
                # 4+
                ws_.cell(row=j, column=ws_.max_column-24).value = \
                    '=IFERROR(' + ws_.cell(row=j, column=ws_.max_column-27).column_letter + str(j) + '+' + ws_.cell(row=j, column=ws_.max_column-26).column_letter + str(j) + ',"--")'


                ws_.cell(row=j, column=ws_.max_column - 0).number_format = '#,##0.00'
                ws_.cell(row=j, column=ws_.max_column - 1).number_format = '#,##0.00'
                ws_.cell(row=j, column=ws_.max_column - 2).number_format = '#,##0.00'
                ws_.cell(row=j, column=ws_.max_column - 3).number_format = '#,##0.00'
                ws_.cell(row=j, column=ws_.max_column - 4).number_format = '#,##0.00'
                ws_.cell(row=j, column=ws_.max_column - 5).number_format = '#,##0.00'
                ws_.cell(row=j, column=ws_.max_column - 6).number_format = '#,##0.00'
                ws_.cell(row=j, column=ws_.max_column - 7).number_format = '#,##0.00'
                ws_.cell(row=j, column=ws_.max_column - 8).number_format = '#,##0.00'
                ws_.cell(row=j, column=ws_.max_column - 9).number_format = '#,##0.00'
                ws_.cell(row=j, column=ws_.max_column - 10).number_format = '#,##0.00'
                ws_.cell(row=j, column=ws_.max_column - 11).number_format = '#,##0.00'
                ws_.cell(row=j, column=ws_.max_column - 12).number_format = '#,##0.00'
                ws_.cell(row=j, column=ws_.max_column - 13).number_format = '#,##0.00'
                ws_.cell(row=j, column=ws_.max_column - 14).number_format = '#,##0.00'

                if(j==3):
                    # Last 5 SR
                    ws_.cell(row=j, column=ws_.max_column - 8).value = \
                        self.__find_bowler_5_matches_sr_(URL_BOWLER_INNINGS_OVERALL_STATS_)
                    # Last 2nd 0
                    ws_.cell(row=j, column=ws_.max_column-16).value = \
                        self.__find_bowler_2nd_last_match_(-1, 1, URL_BOWLER_INNINGS_OVERALL_STATS_)
                    # Last 0
                    ws_.cell(row=j, column=ws_.max_column-17).value = \
                        self.__find_bowler_last_match_(-1, 1, URL_BOWLER_INNINGS_OVERALL_STATS_)
                    # Last 2nd 2+
                    ws_.cell(row=j, column=ws_.max_column-18).value = \
                        self.__find_bowler_2nd_last_match_(1, 4, URL_BOWLER_INNINGS_OVERALL_STATS_)
                    # Last 2+
                    ws_.cell(row=j, column=ws_.max_column-19).value = \
                        self.__find_bowler_last_match_(1, 4, URL_BOWLER_INNINGS_OVERALL_STATS_)
                    # Last 2nd 4+
                    ws_.cell(row=j, column=ws_.max_column-20).value = \
                        self.__find_bowler_2nd_last_match_(3, 11, URL_BOWLER_INNINGS_OVERALL_STATS_)
                    # Last 4+
                    ws_.cell(row=j, column=ws_.max_column-21).value = \
                        self.__find_bowler_last_match_(3, 11, URL_BOWLER_INNINGS_OVERALL_STATS_)
                    # 0
                    ws_.cell(row=j, column=ws_.max_column - 22).value = \
                        self.__find_bowler_dismissals_matches_(-1, 1, URL_BOWLER_INNINGS_OVERALL_STATS_)
                    # 2+
                    ws_.cell(row=j, column=ws_.max_column - 23).value = \
                        self.__find_bowler_dismissals_matches_(1, 4, URL_BOWLER_INNINGS_OVERALL_STATS_)


                if(j==4):
                    # Last 5 SR
                    ws_.cell(row=j, column=ws_.max_column - 8).value = \
                        self.__find_bowler_5_matches_sr_(URL_BOWLER_INNINGS_BAT_1ST_STATS_)
                    # Last 2nd 0
                    ws_.cell(row=j, column=ws_.max_column-16).value = \
                        self.__find_bowler_2nd_last_match_(-1, 1, URL_BOWLER_INNINGS_BAT_1ST_STATS_)
                    # Last 0
                    ws_.cell(row=j, column=ws_.max_column-17).value = \
                        self.__find_bowler_last_match_(-1, 1, URL_BOWLER_INNINGS_BAT_1ST_STATS_)
                    # Last 2nd 2+
                    ws_.cell(row=j, column=ws_.max_column-18).value = \
                        self.__find_bowler_2nd_last_match_(1, 4, URL_BOWLER_INNINGS_BAT_1ST_STATS_)
                    # Last 2+
                    ws_.cell(row=j, column=ws_.max_column-19).value = \
                        self.__find_bowler_last_match_(1, 4, URL_BOWLER_INNINGS_BAT_1ST_STATS_)
                    # Last 2nd 4+
                    ws_.cell(row=j, column=ws_.max_column-20).value = \
                        self.__find_bowler_2nd_last_match_(3, 11, URL_BOWLER_INNINGS_BAT_1ST_STATS_)
                    # Last 4+
                    ws_.cell(row=j, column=ws_.max_column-21).value = \
                        self.__find_bowler_last_match_(3, 11, URL_BOWLER_INNINGS_BAT_1ST_STATS_)
                    # 0
                    ws_.cell(row=j, column=ws_.max_column - 22).value = \
                        self.__find_bowler_dismissals_matches_(-1, 1, URL_BOWLER_INNINGS_BAT_1ST_STATS_)
                    # 2+
                    ws_.cell(row=j, column=ws_.max_column - 23).value = \
                        self.__find_bowler_dismissals_matches_(1, 4, URL_BOWLER_INNINGS_BAT_1ST_STATS_)

                if(j==5):
                    # Last 5 SR
                    ws_.cell(row=j, column=ws_.max_column - 8).value = \
                        self.__find_bowler_5_matches_sr_(URL_BOWLER_INNINGS_BAT_2ND_STATS_)
                    # Last 2nd 0
                    ws_.cell(row=j, column=ws_.max_column-16).value = \
                        self.__find_bowler_2nd_last_match_(-1, 1, URL_BOWLER_INNINGS_BAT_2ND_STATS_)
                    # Last 0
                    ws_.cell(row=j, column=ws_.max_column-17).value = \
                        self.__find_bowler_last_match_(-1, 1, URL_BOWLER_INNINGS_BAT_2ND_STATS_)
                    # Last 2nd 2+
                    ws_.cell(row=j, column=ws_.max_column-18).value = \
                        self.__find_bowler_2nd_last_match_(1, 4, URL_BOWLER_INNINGS_BAT_2ND_STATS_)
                    # Last 2+
                    ws_.cell(row=j, column=ws_.max_column-19).value = \
                        self.__find_bowler_last_match_(1, 4, URL_BOWLER_INNINGS_BAT_2ND_STATS_)
                    # Last 2nd 4+
                    ws_.cell(row=j, column=ws_.max_column-20).value = \
                        self.__find_bowler_2nd_last_match_(3, 11, URL_BOWLER_INNINGS_BAT_2ND_STATS_)
                    # Last 4+
                    ws_.cell(row=j, column=ws_.max_column-21).value = \
                        self.__find_bowler_last_match_(3, 11, URL_BOWLER_INNINGS_BAT_2ND_STATS_)
                    # 0
                    ws_.cell(row=j, column=ws_.max_column - 22).value = \
                        self.__find_bowler_dismissals_matches_(-1, 1, URL_BOWLER_INNINGS_BAT_2ND_STATS_)
                    # 2+
                    ws_.cell(row=j, column=ws_.max_column - 23).value = \
                        self.__find_bowler_dismissals_matches_(1, 4, URL_BOWLER_INNINGS_BAT_2ND_STATS_)

                if(j==10):
                    # Last 5 SR
                    ws_.cell(row=j, column=ws_.max_column - 8).value = \
                        self.__find_bowler_5_matches_sr_(URL_BOWLER_POS_INNINGS_OVERALL_STATS_)
                    # Last 2nd 0
                    ws_.cell(row=j, column=ws_.max_column-16).value = \
                        self.__find_bowler_2nd_last_match_(-1, 1, URL_BOWLER_POS_INNINGS_OVERALL_STATS_)
                    # Last 0
                    ws_.cell(row=j, column=ws_.max_column-17).value = \
                        self.__find_bowler_last_match_(-1, 1, URL_BOWLER_POS_INNINGS_OVERALL_STATS_)
                    # Last 2nd 2+
                    ws_.cell(row=j, column=ws_.max_column-18).value = \
                        self.__find_bowler_2nd_last_match_(1, 4, URL_BOWLER_POS_INNINGS_OVERALL_STATS_)
                    # Last 2+
                    ws_.cell(row=j, column=ws_.max_column-19).value = \
                        self.__find_bowler_last_match_(1, 4, URL_BOWLER_POS_INNINGS_OVERALL_STATS_)
                    # Last 2nd 4+
                    ws_.cell(row=j, column=ws_.max_column-20).value = \
                        self.__find_bowler_2nd_last_match_(3, 11, URL_BOWLER_POS_INNINGS_OVERALL_STATS_)
                    # Last 4+
                    ws_.cell(row=j, column=ws_.max_column-21).value = \
                        self.__find_bowler_last_match_(3, 11, URL_BOWLER_POS_INNINGS_OVERALL_STATS_)
                    # 0
                    ws_.cell(row=j, column=ws_.max_column - 22).value = \
                        self.__find_bowler_dismissals_matches_(-1, 1, URL_BOWLER_POS_INNINGS_OVERALL_STATS_)
                    # 2+
                    ws_.cell(row=j, column=ws_.max_column - 23).value = \
                        self.__find_bowler_dismissals_matches_(1, 4, URL_BOWLER_POS_INNINGS_OVERALL_STATS_)

            j = j + 1

        ws_['A1'].value = player_role
        ws_.freeze_panes = ws_.cell(row=3, column=ws_.max_column-25)
        self.bolding_header_font(ws_)
        self.align_center(ws_)
        # self.adjust_column_width(ws_)

    def __get_recent_bowling_position_(self, url):
        df_list_ = self.Pnds.read_html(url)
        df_stats_ = self.DFrame(df_list_[3])
        df_stats_total_rows_ = len(df_stats_.index)
        found_bowling_pos = False
        for i in range(df_stats_total_rows_ - 1):
            pos = ''
            try:
                pos = df_stats_.at[(df_stats_total_rows_ - 1) - i, 'Pos']
            except ValueError:
                pass
            try:
                pos = int(pos)
                found_bowling_pos = True
                break
            except ValueError:
                pass
        if (not found_bowling_pos):
            pos = '-'
        return pos

    def __get_recent_bowling_position_df_stats_(self, df_stats, bowling_position):
        df_stats = self.DFrame(df_stats)
        if (bowling_position == 1):
            bowling_position = '1st position'
        elif (bowling_position == 2):
            bowling_position = '2nd position'
        elif (bowling_position == 3):
            bowling_position = '3rd position'
        else:
            bowling_position = str(bowling_position) + 'th position'

        print(bowling_position)
        df_recent_bowling_position_stats = df_stats.loc[[bowling_position]]
        return df_recent_bowling_position_stats

    def __find_bowler_5_matches_sr_(self, url):
        df_list_ = self.Pnds.read_html(url)
        df_stats_ = self.DFrame(df_list_[3])
        df_stats_total_rows_ = len(df_stats_.index)
        found_last_5th_match = False
        last_match = 0
        last_5_matches_balls = 0
        last_5_matches_wkts = 0
        last_5_matches_sr = 0
        for i in range(df_stats_total_rows_ - 1):
            overs = ''
            wkts = ''
            try:
                overs = df_stats_.at[(df_stats_total_rows_ - 1) - i, 'Overs']
                wkts = df_stats_.at[(df_stats_total_rows_ - 1) - i, 'Wkts']
            except ValueError:
                pass
            try:
                float(overs)
                last_match = last_match + 1
            except:
                pass
            try:
                overs = float(overs)
                wkts = int(wkts)
                balls = overs * 6
                last_5_matches_wkts = last_5_matches_wkts + wkts
                last_5_matches_balls = last_5_matches_balls + balls
                if (last_match == 5):
                    found_last_5th_match = True
                    try:
                        last_5_matches_sr = last_5_matches_balls / last_5_matches_wkts
                    except ZeroDivisionError:
                        last_5_matches_sr = '--'
                    break
            except ValueError:
                pass
        if (not found_last_5th_match):
            last_5_matches_sr = '-'
        return last_5_matches_sr

    def __find_bowler_last_match_(self, greater_than, less_than, bowler_innings_url):
        df_list_ = self.Pnds.read_html(bowler_innings_url)
        df_stats_ = self.DFrame(df_list_[3])
        df_stats_total_rows_ = len(df_stats_.index)
        j = 0
        found_last_match = False
        for i in range(df_stats_total_rows_-1):
            val = ''
            try:
                val = df_stats_.at[(df_stats_total_rows_-1) - i, 'Wkts']
            except ValueError:
                pass
            try:
                val = int(val)
                j = i
                if (val > greater_than and val < less_than):
                    found_last_match = True
                    break
            except ValueError:
                    pass
        if(not found_last_match):
            j = '-'
        return j

    def __find_bowler_2nd_last_match_(self, greater_than, less_than, bowler_innings_url):
        df_list_ = self.Pnds.read_html(bowler_innings_url)
        df_stats_ = self.DFrame(df_list_[3])
        df_stats_total_rows_ = len(df_stats_.index)
        j = 0
        found_2nd_last_match = False
        last_match = 0
        k = 0
        for i in range(df_stats_total_rows_ - 1):
            val = ''
            try:
                val = df_stats_.at[(df_stats_total_rows_ - 1) - i, 'Wkts']
            except ValueError:
                pass
            try:
                val = int(val)
                j = i
                if (val > greater_than and val < less_than):
                    last_match = last_match + 1
                    if(last_match==1):
                        k = j
                    if(last_match==2):
                        found_2nd_last_match = True
                        j = j - k
                        break
            except ValueError as e:
                pass
        if (not found_2nd_last_match):
            j = '-'
        return j

    def __find_bowler_dismissals_matches_(self, greater_than, less_than, bowler_innings_url):
        df_list_ = self.Pnds.read_html(bowler_innings_url)
        df_stats_ = self.DFrame(df_list_[3])
        df_stats_total_rows_ = len(df_stats_.index)
        total_dismissal_matches = 0
        found_single_match = False
        for i in range(df_stats_total_rows_-1):
            val = ''
            try:
                val = df_stats_.at[(df_stats_total_rows_-1) - i, 'Wkts']
            except ValueError:
                pass
            try:
                val = int(val)
                if (val > greater_than and val < less_than):
                    found_single_match = True
                    total_dismissal_matches = total_dismissal_matches + 1
            except ValueError:
                    pass
        if(not found_single_match):
            total_dismissal_matches = '-'
        return total_dismissal_matches


    def __excel_ready_df_stats_(self, df_stats_, ws_):
        excel_ready_df_stats_ = self.Odf.dataframe_to_rows(df_stats_, index=True, header=True)
        for df_row in excel_ready_df_stats_:
            ws_.append(df_row)

    def bolding_header_font(self, ws):
        for i in range(ws.max_column + 1):
            if (i > 0):
                cell = ws.cell(row=1, column=i)
                cell.font = self.Ostyles.Font(bold=True)

    def align_center(self, ws):
        for i in range(ws.max_column + 1):
            if (i > 0):
                for j in range(ws.max_row + 1):
                    if (j > 0):
                        cell = ws.cell(row=j, column=i)
                        cell.alignment = self.Ostyles.Alignment(horizontal='center', vertical='center')

    def adjust_column_width(self, ws):
        for col in ws.columns:
            max_lenght = 0
            col_name = self.Re.findall('\w\d', str(col[0]))
            col_name = col_name[0]
            col_name = self.Re.findall('\w', str(col_name))[0]
            for cell in col:
                try:
                    if len(str(cell.value)) > max_lenght:
                        max_lenght = len(cell.value)
                except:
                    pass
            adjusted_width = (max_lenght + 2)
            ws.column_dimensions[col_name].width = adjusted_width
