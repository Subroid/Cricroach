
import urllib.request as __urlreq
from  bs4 import BeautifulSoup as __soup
import re
from pandas import DataFrame as __df
from openpyxl import Workbook as __wb
import openpyxl.utils.dataframe as __odf
from openpyxl import load_workbook
import openpyxl.styles as __oxl_styles

_url = __urlreq.urlopen("http://www.espncricinfo.com/series/8039/game/1144483/england-vs-south-africa-1st-match-icc-cricket-world-cup-2019")
_squads = ['England', 'South Africa']
_excel_file = "C:/Users/ASUS/Desktop/Next Match Squads.xlsx"
_new_wb = True

_parser = 'html.parser'
_soup = __soup(_url, _parser)
_table_tag_list = _soup.find_all('table')

_player_name_list = list()
_player_role_list = list()
_player_id_list = list()

_table_tag = _table_tag_list[0]
_player_label = _table_tag.th.string
_tr_tag_list = _table_tag.find_all('tr')
_length_tr_tags = len(_tr_tag_list)

def parse_html_scrpe_data():
    for i in range(_length_tr_tags):
        if(i > 0):
            tr_tag = _tr_tag_list[i]
            td_tags = tr_tag.findAll('td')
            td_tag_1 = td_tags[0]
            a_tag = td_tag_1.a
            link = a_tag['href']
            td_tag_2 = td_tags[1]
            player_name = td_tag_1.string.strip()
            player_role = td_tag_2.string.strip()
            player_id_list = re.findall('\d+', link)
            player_id = player_id_list[0]
            _player_name_list.append(player_name)
            _player_role_list.append(player_role)
            _player_id_list.append(player_id)
parse_html_scrpe_data()

_players_df = __df({'Player Name': _player_name_list, 'Player Role': _player_role_list, 'Player ID': _player_id_list})
_wb = __wb()
_ws = _wb.create_sheet(_squads[0])

def appending_rows():
    _deployable_odf = __odf.dataframe_to_rows(_players_df, index=False, header=True)
    for df_row in _deployable_odf:
        _ws.append(df_row)
appending_rows()

def bolding_font():
    for i in range(_ws.max_column+1):
        if(i > 0):
            cell = _ws.cell(row=1,column=i)
            cell.font = __oxl_styles.Font(bold=True)
bolding_font()

def align_center():
    for i in range(_ws.max_column+1):
        if(i > 0):
            for j in range( _ws.max_row+1):
                if(j > 0):
                    cell = _ws.cell(row=j,column=i)
                    cell.alignment = __oxl_styles.Alignment(horizontal='center', vertical='center')
align_center()

def adjust_column_width():
    for col in _ws.columns:
        max_lenght = 0
        col_name = re.findall('\w\d', str(col[0]))
        col_name = col_name[0]
        col_name = re.findall('\w', str(col_name))[0]
        for cell in col:
            try:
                if len(str(cell.value)) > max_lenght:
                    max_lenght = len(cell.value)
            except:
                pass
        adjusted_width = (max_lenght+2)
        _ws.column_dimensions[col_name].width = adjusted_width
adjust_column_width()

_wb.save(_excel_file)


_player_name_list.clear()
_player_role_list.clear()
_player_id_list.clear()

_table_tag = _table_tag_list[1]
_player_label = _table_tag.th.string
_tr_tag_list = _table_tag.find_all('tr')
_length_tr_tags = len(_tr_tag_list)

parse_html_scrpe_data()

_players_df = __df({'Player Name': _player_name_list, 'Player Role': _player_role_list, 'Player ID': _player_id_list})
_wb = load_workbook(_excel_file)
_ws = _wb.create_sheet(title=_squads[1])

appending_rows()

bolding_font()

align_center()

adjust_column_width()

_wb.save(_excel_file)

