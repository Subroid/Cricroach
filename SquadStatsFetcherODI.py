

import urllib.request as url_req__
from bs4 import BeautifulSoup as bsoup__
import pandas as pd__
from openpyxl import Workbook as wb__
import openpyxl.utils.dataframe as odf__
from openpyxl import load_workbook
import re as re__
from bs4 import Tag as TAG
from pandas import DataFrame as DF
from openpyxl.worksheet.worksheet import Worksheet as WS


url_of_match_ = 'http://www.espncricinfo.com/series/8039/game/1144488/england-vs-pakistan-6th-match-icc-cricket-world-cup-2019'
url_main_domain_ = 'http://www.espncricinfo.com/'
url_resp_ = url_req__.urlopen(url_of_match_)
PARSER_ = 'html.parser'

excel_file_loc_ = 'E:/D11/D11/'

bsoup_ = bsoup__(url_resp_, PARSER_)

team_names_span_tags_ = bsoup_.find_all('span', attrs={'class': 'team-name-short'})
team_1_name_ = team_names_span_tags_[0].string
team_2_name_ = team_names_span_tags_[1].string

team_table_tags_ = bsoup_.find_all('table')
team_1_table_tag_ = team_table_tags_[0]
team_2_table_tag_ = team_table_tags_[1]

team_1_tbody_tag_ = team_1_table_tag_.tbody
team_1_tr_tags_ = team_1_tbody_tag_.find_all('tr')


excel_file_ = excel_file_loc_ + team_1_name_ + ' Batting career summary.xlsx'
wb_ = wb__()

for i in range(len(team_1_tr_tags_)):
    player_tr_tag_ = team_1_tr_tags_[i]
    player_a_tag_ = player_tr_tag_.a
    player_profile_link_ = player_a_tag_['href']
    player_id_ = re__.findall('\d+', player_profile_link_)[0]
    player_name_ = str(player_a_tag_.string).strip()
    player_role_td_tag_ = player_tr_tag_.find('td', {'class': 'role'})
    player_role_ = str(player_role_td_tag_.string)

    '''http://stats.espncricinfo.com/ci/engine/player/24598.html?class=2;template=results;type=batting'''
    url_player_stats_ = 'http://stats.espncricinfo.com/ci/engine/player/' + player_id_ + '.html?class=2;template=results;type=batting'
    # url_resp_ = url_req__.urlopen(url_player_stats_)
    # bsoup_ = bsoup__(url_resp_, PARSER_)
    df_list_stats_ = pd__.read_html(url_player_stats_)
    df_stats_1_ = df_list_stats_[2]
    df_stats_2_ = df_list_stats_[3]
    df_stats_2_ = df_stats_2_.dropna(axis=1, how='all')
    df_stats_2_ = df_stats_2_.fillna("")

    ws_ = wb_.create_sheet(player_name_)
    deployable_odf_ = odf__.dataframe_to_rows(df_stats_1_, index=False, header=True)
    for df_row in deployable_odf_:
        ws_.append(df_row)
    ws_.append([""])
    deployable_odf_ = odf__.dataframe_to_rows(df_stats_2_, index=False, header=True)
    for df_row in deployable_odf_:
        ws_.append(df_row)

    ws_.cell(row=1, column=ws_.max_column + 2).value = '0-19'
    ws_.cell(row=1, column=ws_.max_column + 1).value = 'Last 100'
    ws_.cell(row=1, column=ws_.max_column + 1).value = 'Last 50'
    ws_.cell(row=1, column=ws_.max_column + 1).value = 'Last 0-19'
    ws_.cell(row=1, column=ws_.max_column + 1).value = '4s avg'
    ws_.cell(row=1, column=ws_.max_column + 1).value = '6s avg'
    ws_.cell(row=1, column=ws_.max_column + 1).value = '100s avg'
    ws_.cell(row=1, column=ws_.max_column + 1).value = '50s avg'
    ws_.cell(row=1, column=ws_.max_column + 1).value = '0-19 avg'
    ws_.cell(row=1, column=ws_.max_column + 1).value = '100 due'
    ws_.cell(row=1, column=ws_.max_column + 1).value = '50 due'
    ws_.cell(row=1, column=ws_.max_column + 1).value = '0-19 due'

    j = 2;
    for row in ws_.iter_rows(min_row=2, max_row=ws_.max_row):
        ws_.cell(row=j, column=ws_.max_column).value = '=IFERROR(S' + str(j) + '-U' + str(j) + ',"--")'
        ws_.cell(row=j, column=ws_.max_column-1).value = '=IFERROR(Y' + str(j) + '-T' + str(j) + ',"--")'
        ws_.cell(row=j, column=ws_.max_column-2).value = '=IFERROR(X' + str(j) + '-S' + str(j) + ',"--")'
        ws_.cell(row=j, column=ws_.max_column-3).value = '=IFERROR(D' + str(j) + '/R' + str(j) + ',"--")'
        ws_.cell(row=j, column=ws_.max_column-4).value = '=IFERROR(D' + str(j) + '/L' + str(j) + ',"--")'
        ws_.cell(row=j, column=ws_.max_column-5).value = '=IFERROR(D' + str(j) + '/K' + str(j) + ',"--")'
        ws_.cell(row=j, column=ws_.max_column-6).value = '=IFERROR(I'+str(j)+'/O'+str(j)+',"--")'
        ws_.cell(row=j, column=ws_.max_column-3).value = '=IFERROR(I'+str(j)+'/N'+str(j)+',"--")'
        j = j+1

    ws_['A1'].value = player_role_
    #todo freeze panes at B5
    #todo V1 to AC1 italic
    #todo matches batting/fielding first cells bold

sheet_ = wb_['Sheet']
wb_.remove(sheet_)
wb_.save(excel_file_)
wb_.close()

excel_file_ = excel_file_loc_ + team_1_name_ + ' Batting innings by innings.xlsx'
wb_ = wb__()

for i in range(len(team_1_tr_tags_)):
    player_tr_tag_ = team_1_tr_tags_[i]
    player_a_tag_ = player_tr_tag_.a
    player_profile_link_ = player_a_tag_['href']
    player_id_ = re__.findall('\d+', player_profile_link_)[0]
    player_name_ = str(player_a_tag_.string).strip()
    player_role_td_tag_ = player_tr_tag_.find('td', {'class': 'role'})
    player_role_ = str(player_role_td_tag_.string)

    '''http://stats.espncricinfo.com/ci/engine/player/24598.html?class=2;template=results;type=batting;view=innings'''
    url_player_stats_ = 'http://stats.espncricinfo.com/ci/engine/player/' + player_id_ + '.html?class=2;template=results;type=batting;view=innings'

    df_list_stats_ = pd__.read_html(url_player_stats_)
    df_stats_1_ = df_list_stats_[2]
    df_stats_2_ = df_list_stats_[3]

    ws_ = wb_.create_sheet(player_name_)
    deployable_odf_ = odf__.dataframe_to_rows(df_stats_1_, index=False, header=True)
    for df_row in deployable_odf_:
        ws_.append(df_row)
    ws_.append([""])
    deployable_odf_ = odf__.dataframe_to_rows(df_stats_2_, index=False, header=True)
    for df_row in deployable_odf_:
        ws_.append(df_row)
    ws_['A1'].value = player_role_

sheet_ = wb_['Sheet']
wb_.remove(sheet_)
wb_.save(excel_file_)
wb_.close()

excel_file_ = excel_file_loc_ + team_1_name_ + ' Bowling career summary.xlsx'
wb_ = wb__()

for i in range(len(team_1_tr_tags_)):
    player_tr_tag_ = team_1_tr_tags_[i]
    player_a_tag_ = player_tr_tag_.a
    player_profile_link_ = player_a_tag_['href']
    player_id_ = re__.findall('\d+', player_profile_link_)[0]
    player_name_ = str(player_a_tag_.string).strip()
    player_role_td_tag_ = player_tr_tag_.find('td', {'class': 'role'})
    player_role_ = str(player_role_td_tag_.string)

    '''http://stats.espncricinfo.com/ci/engine/player/24598.html?class=2;template=results;type=bowling'''
    url_player_stats_ = 'http://stats.espncricinfo.com/ci/engine/player/' + player_id_ + '.html?class=2;template=results;type=bowling'
    # url_resp_ = url_req__.urlopen(url_player_stats_)
    # bsoup_ = bsoup__(url_resp_, PARSER_)
    df_list_stats_ = pd__.read_html(url_player_stats_)
    df_stats_1_ = df_list_stats_[2]
    df_stats_2_ = df_list_stats_[3]
    df_stats_2_ = df_stats_2_.dropna(axis=1, how='all')
    df_stats_2_ = df_stats_2_.fillna("")

    ws_ = wb_.create_sheet(player_name_)
    deployable_odf_ = odf__.dataframe_to_rows(df_stats_1_, index=False, header=True)
    for df_row in deployable_odf_:
        ws_.append(df_row)
    ws_.append([""])
    deployable_odf_ = odf__.dataframe_to_rows(df_stats_2_, index=False, header=True)
    for df_row in deployable_odf_:
        ws_.append(df_row)

    ws_.cell(row=1, column=ws_.max_column).value = '4+'
    ws_.cell(row=1, column=ws_.max_column + 1).value = 'Last 4+'
    ws_.cell(row=1, column=ws_.max_column + 1).value = 'Last 2+'
    ws_.cell(row=1, column=ws_.max_column + 1).value = 'Last 0+'
    ws_.cell(row=1, column=ws_.max_column + 1).value = '2+'
    ws_.cell(row=1, column=ws_.max_column + 1).value = '0'
    ws_.cell(row=1, column=ws_.max_column + 1).value = '4+ avg'
    ws_.cell(row=1, column=ws_.max_column + 1).value = '2+ avg'
    ws_.cell(row=1, column=ws_.max_column + 1).value = '0 avg'
    ws_.cell(row=1, column=ws_.max_column + 1).value = '4+ due'
    ws_.cell(row=1, column=ws_.max_column + 1).value = '2+ due'
    ws_.cell(row=1, column=ws_.max_column + 1).value = '0 due'
    ws_.cell(row=1, column=ws_.max_column + 1).value = 'Overs avg'

    j = 2;
    for row in ws_.iter_rows(min_row=2, max_row=ws_.max_row):
        #TODO didn't work
        # if(str(ws_.cell(row=j, column=ws_.max_column-1).value) == "" and str(ws_.cell(row=j, column=ws_.max_column-2).value) == ""):
        #     str(ws_.cell(row=j, column=ws_.max_column - 1).value) == "--"
        ws_.cell(row=j, column=ws_.max_column-12).value = '=IFERROR(M'+str(j)+'+N'+str(j)+',"--")'
        if(j==4):
            ws_.cell(row=j, column=ws_.max_column-12).value = ''
        ws_.cell(row=j, column=ws_.max_column-1).value = '=IFERROR(W'+str(j)+'-R'+str(j)+',"--")'
        if(j==4):
            ws_.cell(row=j, column=ws_.max_column-1).value = ''
        ws_.cell(row=j, column=ws_.max_column-2).value = '=IFERROR(V'+str(j)+'-Q'+str(j)+',"--")'
        if(j==4):
            ws_.cell(row=j, column=ws_.max_column-2).value = ''
        ws_.cell(row=j, column=ws_.max_column-3).value = '=IFERROR(U'+str(j)+'-P'+str(j)+',"--")'
        if(j==4):
            ws_.cell(row=j, column=ws_.max_column-3).value = ''
        ws_.cell(row=j, column=ws_.max_column-4).value = '=IFERROR(D'+str(j)+'/T'+str(j)+',"--")'
        if(j==4):
            ws_.cell(row=j, column=ws_.max_column-4).value = ''
        ws_.cell(row=j, column=ws_.max_column-5).value = '=IFERROR(D'+str(j)+'/S'+str(j)+',"--")'
        if(j==4):
            ws_.cell(row=j, column=ws_.max_column-5).value = ''
        ws_.cell(row=j, column=ws_.max_column-6).value = '=IFERROR(D'+str(j)+'/O'+str(j)+',"--")'
        if(j==4):
            ws_.cell(row=j, column=ws_.max_column-6).value = ''
        ws_.cell(row=j, column=ws_.max_column).value = '=IFERROR(E' + str(j) + '/C' + str(j) + ',"--")'
        if (j == 4):
            ws_.cell(row=j, column=ws_.max_column).value = ''
        j = j+1
    ws_['A1'].value = player_role_

sheet_ = wb_['Sheet']
wb_.remove(sheet_)
wb_.save(excel_file_)
wb_.close()

excel_file_ = excel_file_loc_ + team_1_name_ + ' Bowling innings by innings.xlsx'
wb_ = wb__()

for i in range(len(team_1_tr_tags_)):
    player_tr_tag_ = team_1_tr_tags_[i]
    player_a_tag_ = player_tr_tag_.a
    player_profile_link_ = player_a_tag_['href']
    player_id_ = re__.findall('\d+', player_profile_link_)[0]
    player_name_ = str(player_a_tag_.string).strip()
    player_role_td_tag_ = player_tr_tag_.find('td', {'class': 'role'})
    player_role_ = str(player_role_td_tag_.string)

    '''http://stats.espncricinfo.com/ci/engine/player/24598.html?class=2;template=results;type=bowling;view=innings'''
    url_player_stats_ = 'http://stats.espncricinfo.com/ci/engine/player/' + player_id_ + '.html?class=2;template=results;type=bowling;view=innings'

    df_list_stats_ = pd__.read_html(url_player_stats_)
    df_stats_1_ = df_list_stats_[2]
    df_stats_2_ = df_list_stats_[3]

    ws_ = wb_.create_sheet(player_name_)
    deployable_odf_ = odf__.dataframe_to_rows(df_stats_1_, index=False, header=True)
    for df_row in deployable_odf_:
        ws_.append(df_row)
    ws_.append([""])
    deployable_odf_ = odf__.dataframe_to_rows(df_stats_2_, index=False, header=True)
    for df_row in deployable_odf_:
        ws_.append(df_row)
    ws_['A1'].value = player_role_

sheet_ = wb_['Sheet']
wb_.remove(sheet_)
wb_.save(excel_file_)
wb_.close()


team_2_tbody_tag_ = team_2_table_tag_.tbody
team_2_tr_tags_ = team_2_tbody_tag_.find_all('tr')


excel_file_ = excel_file_loc_ + team_2_name_ + ' Batting career summary.xlsx'
wb_ = wb__()

for i in range(len(team_2_tr_tags_)):
    player_tr_tag_ = team_2_tr_tags_[i]
    player_a_tag_ = player_tr_tag_.a
    player_profile_link_ = player_a_tag_['href']
    player_id_ = re__.findall('\d+', player_profile_link_)[0]
    player_name_ = str(player_a_tag_.string).strip()
    player_role_td_tag_ = player_tr_tag_.find('td', {'class': 'role'})
    player_role_ = str(player_role_td_tag_.string)

    '''http://stats.espncricinfo.com/ci/engine/player/24598.html?class=2;template=results;type=batting'''
    url_player_stats_ = 'http://stats.espncricinfo.com/ci/engine/player/' + player_id_ + '.html?class=2;template=results;type=batting'
    # url_resp_ = url_req__.urlopen(url_player_stats_)
    # bsoup_ = bsoup__(url_resp_, PARSER_)
    df_list_stats_ = pd__.read_html(url_player_stats_)
    df_stats_1_ = df_list_stats_[2]
    df_stats_2_ = df_list_stats_[3]
    df_stats_2_ = df_stats_2_.dropna(axis=1, how='all')
    df_stats_2_ = df_stats_2_.fillna("")

    ws_ = wb_.create_sheet(player_name_)
    deployable_odf_ = odf__.dataframe_to_rows(df_stats_1_, index=False, header=True)
    for df_row in deployable_odf_:
        ws_.append(df_row)
    ws_.append([""])
    deployable_odf_ = odf__.dataframe_to_rows(df_stats_2_, index=False, header=True)
    for df_row in deployable_odf_:
        ws_.append(df_row)

    ws_.cell(row=1, column=ws_.max_column + 2).value = '4s avg'
    ws_.cell(row=1, column=ws_.max_column + 1).value = '6s avg'
    ws_.cell(row=1, column=ws_.max_column + 1).value = '100s avg'
    ws_.cell(row=1, column=ws_.max_column + 1).value = '50s avg'

    j = 2;
    for row in ws_.iter_rows(min_row=2, max_row=ws_.max_row):
        ws_.cell(row=j, column=ws_.max_column-3).value = '=IFERROR(I'+str(j)+'/N'+str(j)+',"--")'
        ws_.cell(row=j, column=ws_.max_column-2).value = '=IFERROR(I'+str(j)+'/O'+str(j)+',"--")'
        ws_.cell(row=j, column=ws_.max_column-1).value = '=IFERROR(D'+str(j)+'/K'+str(j)+',"--")'
        ws_.cell(row=j, column=ws_.max_column).value = '=IFERROR(D'+str(j)+'/L'+str(j)+',"--")'
        j = j+1
    ws_['A1'].value = player_role_

sheet_ = wb_['Sheet']
wb_.remove(sheet_)
wb_.save(excel_file_)
wb_.close()

excel_file_ = excel_file_loc_ + team_2_name_ + ' Batting innings by innings.xlsx'
wb_ = wb__()

for i in range(len(team_2_tr_tags_)):
    player_tr_tag_ = team_2_tr_tags_[i]
    player_a_tag_ = player_tr_tag_.a
    player_profile_link_ = player_a_tag_['href']
    player_id_ = re__.findall('\d+', player_profile_link_)[0]
    player_name_ = str(player_a_tag_.string).strip()
    player_role_td_tag_ = player_tr_tag_.find('td', {'class': 'role'})
    player_role_ = str(player_role_td_tag_.string)

    '''http://stats.espncricinfo.com/ci/engine/player/24598.html?class=2;template=results;type=batting;view=innings'''
    url_player_stats_ = 'http://stats.espncricinfo.com/ci/engine/player/' + player_id_ + '.html?class=2;template=results;type=batting;view=innings'

    df_list_stats_ = pd__.read_html(url_player_stats_)
    df_stats_1_ = df_list_stats_[2]
    df_stats_2_ = df_list_stats_[3]
    df_stats_2_ = df_stats_2_.set_index('Grouping')
    # todo finding batting position and appending to the player_name
    # todo matches batting/fielding first data inseriting to stats_1
    # todo 0-19, lasts, overall/bat-1st-2nd data retrieving and inseritng to stats_1 or excel


    ws_ = wb_.create_sheet(player_name_)
    deployable_odf_ = odf__.dataframe_to_rows(df_stats_1_, index=False, header=True)
    for df_row in deployable_odf_:
        ws_.append(df_row)
    ws_.append([""])
    deployable_odf_ = odf__.dataframe_to_rows(df_stats_2_, index=False, header=True)
    for df_row in deployable_odf_:
        ws_.append(df_row)
    ws_['A1'].value = player_role_

sheet_ = wb_['Sheet']
wb_.remove(sheet_)
wb_.save(excel_file_)
wb_.close()

excel_file_ = excel_file_loc_ + team_2_name_ + ' Bowling career summary.xlsx'
wb_ = wb__()

for i in range(len(team_2_tr_tags_)):
    player_tr_tag_ = team_2_tr_tags_[i]
    player_a_tag_ = player_tr_tag_.a
    player_profile_link_ = player_a_tag_['href']
    player_id_ = re__.findall('\d+', player_profile_link_)[0]
    player_name_ = str(player_a_tag_.string).strip()
    player_role_td_tag_ = player_tr_tag_.find('td', {'class': 'role'})
    player_role_ = str(player_role_td_tag_.string)

    '''http://stats.espncricinfo.com/ci/engine/player/24598.html?class=2;template=results;type=bowling'''
    url_player_stats_ = 'http://stats.espncricinfo.com/ci/engine/player/' + player_id_ + '.html?class=2;template=results;type=bowling'
    # url_resp_ = url_req__.urlopen(url_player_stats_)
    # bsoup_ = bsoup__(url_resp_, PARSER_)
    df_list_stats_ = pd__.read_html(url_player_stats_)
    df_stats_1_ = df_list_stats_[2]
    df_stats_2_ = df_list_stats_[3]
    df_stats_2_ = df_stats_2_.dropna(axis=1, how='all')
    df_stats_2_ = df_stats_2_.fillna("")

    ws_ = wb_.create_sheet(player_name_)
    deployable_odf_ = odf__.dataframe_to_rows(df_stats_1_, index=False, header=True)
    for df_row in deployable_odf_:
        ws_.append(df_row)
    ws_.append([""])
    deployable_odf_ = odf__.dataframe_to_rows(df_stats_2_, index=False, header=True)
    for df_row in deployable_odf_:
        ws_.append(df_row)

    ws_.cell(row=1, column=ws_.max_column).value = '4+'
    ws_.cell(row=1, column=ws_.max_column + 1).value = 'Last 4+'
    ws_.cell(row=1, column=ws_.max_column + 1).value = 'Last 2+'
    ws_.cell(row=1, column=ws_.max_column + 1).value = 'Last 0+'
    ws_.cell(row=1, column=ws_.max_column + 1).value = '2+'
    ws_.cell(row=1, column=ws_.max_column + 1).value = '0'
    ws_.cell(row=1, column=ws_.max_column + 1).value = '4+ avg'
    ws_.cell(row=1, column=ws_.max_column + 1).value = '2+ avg'
    ws_.cell(row=1, column=ws_.max_column + 1).value = '0 avg'
    ws_.cell(row=1, column=ws_.max_column + 1).value = '4+ due'
    ws_.cell(row=1, column=ws_.max_column + 1).value = '2+ due'
    ws_.cell(row=1, column=ws_.max_column + 1).value = '0 due'
    ws_.cell(row=1, column=ws_.max_column + 1).value = 'Overs avg'

    j = 2;
    for row in ws_.iter_rows(min_row=2, max_row=ws_.max_row):
        #TODO didn't work
        # if(str(ws_.cell(row=j, column=ws_.max_column-1).value) == "" and str(ws_.cell(row=j, column=ws_.max_column-2).value) == ""):
        #     str(ws_.cell(row=j, column=ws_.max_column - 1).value) == "--"
        ws_.cell(row=j, column=ws_.max_column - 12).value = '=IFERROR(M' + str(j) + '+N' + str(j) + ',"--")'
        if (j == 4):
            ws_.cell(row=j, column=ws_.max_column - 12).value = ''
        ws_.cell(row=j, column=ws_.max_column - 1).value = '=IFERROR(W' + str(j) + '-R' + str(j) + ',"--")'
        if (j == 4):
            ws_.cell(row=j, column=ws_.max_column - 1).value = ''
        ws_.cell(row=j, column=ws_.max_column - 2).value = '=IFERROR(V' + str(j) + '-Q' + str(j) + ',"--")'
        if (j == 4):
            ws_.cell(row=j, column=ws_.max_column - 2).value = ''
        ws_.cell(row=j, column=ws_.max_column - 3).value = '=IFERROR(U' + str(j) + '-P' + str(j) + ',"--")'
        if (j == 4):
            ws_.cell(row=j, column=ws_.max_column - 3).value = ''
        ws_.cell(row=j, column=ws_.max_column - 4).value = '=IFERROR(D' + str(j) + '/T' + str(j) + ',"--")'
        if (j == 4):
            ws_.cell(row=j, column=ws_.max_column - 4).value = ''
        ws_.cell(row=j, column=ws_.max_column - 5).value = '=IFERROR(D' + str(j) + '/S' + str(j) + ',"--")'
        if (j == 4):
            ws_.cell(row=j, column=ws_.max_column - 5).value = ''
        ws_.cell(row=j, column=ws_.max_column - 6).value = '=IFERROR(D' + str(j) + '/O' + str(j) + ',"--")'
        if (j == 4):
            ws_.cell(row=j, column=ws_.max_column - 6).value = ''
        ws_.cell(row=j, column=ws_.max_column).value = '=IFERROR(E' + str(j) + '/C' + str(j) + ',"--")'
        if (j == 4):
            ws_.cell(row=j, column=ws_.max_column).value = ''
        j = j + 1

sheet_ = wb_['Sheet']
wb_.remove(sheet_)
wb_.save(excel_file_)
wb_.close()

excel_file_ = excel_file_loc_ + team_2_name_ + ' Bowling innings by innings.xlsx'
wb_ = wb__()

for i in range(len(team_2_tr_tags_)):
    player_tr_tag_ = team_2_tr_tags_[i]
    player_a_tag_ = player_tr_tag_.a
    player_profile_link_ = player_a_tag_['href']
    player_id_ = re__.findall('\d+', player_profile_link_)[0]
    player_name_ = str(player_a_tag_.string).strip()
    player_role_td_tag_ = player_tr_tag_.find('td', {'class': 'role'})
    player_role_ = str(player_role_td_tag_.string)

    '''http://stats.espncricinfo.com/ci/engine/player/24598.html?class=2;template=results;type=bowling;view=innings'''
    url_player_stats_ = 'http://stats.espncricinfo.com/ci/engine/player/' + player_id_ + '.html?class=2;template=results;type=bowling;view=innings'

    df_list_stats_ = pd__.read_html(url_player_stats_)
    df_stats_1_ = df_list_stats_[2]
    df_stats_2_ = df_list_stats_[3]

    ws_ = wb_.create_sheet(player_name_)
    deployable_odf_ = odf__.dataframe_to_rows(df_stats_1_, index=False, header=True)
    for df_row in deployable_odf_:
        ws_.append(df_row)
    ws_.append([""])
    deployable_odf_ = odf__.dataframe_to_rows(df_stats_2_, index=False, header=True)
    for df_row in deployable_odf_:
        ws_.append(df_row)
    ws_['A1'].value = player_role_

sheet_ = wb_['Sheet']
wb_.remove(sheet_)
wb_.save(excel_file_)
wb_.close()


